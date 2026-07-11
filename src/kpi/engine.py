"""Calculate annual and current financial KPIs from normalized FOS data."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook

from src.extract import HistoricalExtractionResult
from src.transform import CategoryRegistry

ZERO = Decimal("0")
ONE_HUNDRED = Decimal("100")


@dataclass(frozen=True, slots=True)
class AnnualKPI:
    year: int
    coverage_status: str
    pay_periods: int
    true_income: Decimal
    cash_flow_adjustments: Decimal
    fixed_expenses: Decimal
    variable_irregular_expenses: Decimal
    known_operating_expenses: Decimal
    core_spending: Decimal
    lifestyle_spending: Decimal
    wealth_building: Decimal
    targeted_debt_reduction: Decimal
    excluded_transfers: Decimal
    unknown_amount: Decimal
    fixed_cost_ratio: Decimal | None
    known_expense_ratio: Decimal | None
    wealth_building_rate: Decimal | None
    savings_velocity: Decimal | None
    financial_flexibility: Decimal | None
    housing_ratio: Decimal | None
    transportation_ratio: Decimal | None
    food_ratio: Decimal | None
    lifestyle_ratio: Decimal | None
    data_coverage_ratio: Decimal | None
    comparison_eligible: bool


@dataclass(frozen=True, slots=True)
class CurrentSnapshot:
    source_sheet: str
    latest_complete_year: int
    total_assets: Decimal
    total_liabilities: Decimal
    net_worth: Decimal
    financial_assets: Decimal
    savings_cash: Decimal
    home_equity: Decimal
    mortgage_balance: Decimal
    line_of_credit_balance: Decimal
    vehicle_loan_balance: Decimal
    emergency_fund_months: Decimal | None
    unsecured_debt_ratio: Decimal | None
    fpi_score: Decimal | None
    fpi_band: str


class KPIEngine:
    """Produce FOS v0.4.1 KPIs using only existing workbook data."""

    TARGETED_DEBT_IDS = frozenset({"TRF006", "FIN001"})
    EXCLUDED_TRANSFER_IDS = frozenset({"TRF005", "TRF007", "TRF008"})
    TRUE_INCOME_IDS = frozenset({"INC001", "INC002", "INC003"})
    ADJUSTMENT_IDS = frozenset({"INC004"})
    FINANCIAL_ASSET_PREFIXES = ("TFSA", "RRSP", "RESP", "Savings")

    def __init__(self, category_registry: CategoryRegistry) -> None:
        self.registry = category_registry
        self.entries = {
            str(entry["category_id"]): entry for entry in category_registry.entries()
        }

    @staticmethod
    def _ratio(numerator: Decimal, denominator: Decimal) -> Decimal | None:
        if denominator == 0:
            return None
        return numerator / denominator

    @staticmethod
    def _sum_amount(records) -> Decimal:
        return sum((record.amount for record in records), ZERO)

    def calculate_annual(self, extraction: HistoricalExtractionResult) -> tuple[AnnualKPI, ...]:
        results: list[AnnualKPI] = []
        for sheet in extraction.sheets:
            source = sheet.result
            transactions = (
                list(source.transfers)
                + list(source.variable_expenses)
                + list(source.fixed_expenses)
            )
            true_income = self._sum_amount(
                item for item in source.income if item.category_id in self.TRUE_INCOME_IDS
            )
            adjustments = self._sum_amount(
                item for item in source.income if item.category_id in self.ADJUSTMENT_IDS
            )
            fixed_expenses = self._sum_amount(
                item
                for item in transactions
                if self.entries[item.category_id]["category_type"] == "Fixed Expense"
            )
            variable_irregular = self._sum_amount(
                item
                for item in transactions
                if self.entries[item.category_id]["category_type"]
                in {"Variable Expense", "Irregular Expense", "Capital"}
            )
            known_operating = fixed_expenses + variable_irregular
            core_spending = self._sum_amount(
                item
                for item in transactions
                if self.entries[item.category_id]["financial_purpose"]
                in {"Essential Living", "Family", "Debt Cost"}
            )
            lifestyle = self._sum_amount(
                item
                for item in transactions
                if self.entries[item.category_id]["financial_purpose"] == "Lifestyle"
            )
            wealth = self._sum_amount(
                item
                for item in transactions
                if self.entries[item.category_id]["financial_purpose"] == "Wealth Building"
            )
            targeted_debt = self._sum_amount(
                item for item in transactions if item.category_id in self.TARGETED_DEBT_IDS
            )
            excluded_transfers = self._sum_amount(
                item for item in transactions if item.category_id in self.EXCLUDED_TRANSFER_IDS
            )
            unknown_amount = self._sum_amount(source.unknown_categories)

            by_master: dict[str, Decimal] = {}
            for item in transactions:
                entry = self.entries[item.category_id]
                if entry["category_type"] == "Transfer":
                    continue
                master = str(entry["master_category"])
                by_master[master] = by_master.get(master, ZERO) + item.amount

            fixed_ratio = self._ratio(fixed_expenses, true_income)
            known_ratio = self._ratio(known_operating, true_income)
            wealth_rate = self._ratio(wealth, true_income)
            savings_velocity = self._ratio(wealth + targeted_debt, true_income)
            flexibility = self._ratio(
                true_income - fixed_expenses - targeted_debt, true_income
            )
            mapped_outflows = self._sum_amount(transactions)
            coverage = self._ratio(mapped_outflows, mapped_outflows + unknown_amount)
            pay_periods = len(source.periods)
            coverage_status = "Complete" if pay_periods >= 24 else "Partial"
            comparison_eligible = bool(
                coverage_status == "Complete"
                and coverage is not None
                and coverage >= Decimal("0.85")
            )

            results.append(
                AnnualKPI(
                    year=sheet.year,
                    coverage_status=coverage_status,
                    pay_periods=pay_periods,
                    true_income=true_income,
                    cash_flow_adjustments=adjustments,
                    fixed_expenses=fixed_expenses,
                    variable_irregular_expenses=variable_irregular,
                    known_operating_expenses=known_operating,
                    core_spending=core_spending,
                    lifestyle_spending=lifestyle,
                    wealth_building=wealth,
                    targeted_debt_reduction=targeted_debt,
                    excluded_transfers=excluded_transfers,
                    unknown_amount=unknown_amount,
                    fixed_cost_ratio=fixed_ratio,
                    known_expense_ratio=known_ratio,
                    wealth_building_rate=wealth_rate,
                    savings_velocity=savings_velocity,
                    financial_flexibility=flexibility,
                    housing_ratio=self._ratio(by_master.get("Housing", ZERO), true_income),
                    transportation_ratio=self._ratio(
                        by_master.get("Transportation", ZERO), true_income
                    ),
                    food_ratio=self._ratio(by_master.get("Food", ZERO), true_income),
                    lifestyle_ratio=self._ratio(lifestyle, true_income),
                    data_coverage_ratio=coverage,
                    comparison_eligible=comparison_eligible,
                )
            )
        return tuple(results)

    def extract_balance_sheet(self, workbook_path: str | Path) -> tuple[dict[str, Decimal], dict[str, str]]:
        workbook = load_workbook(workbook_path, data_only=True, read_only=False)
        try:
            if "A & L" not in workbook.sheetnames:
                raise ValueError("Source workbook does not contain the 'A & L' sheet.")
            worksheet = workbook["A & L"]
            amounts: dict[str, Decimal] = {}
            classifications: dict[str, str] = {}
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                item, classification, amount, _comments = (list(row) + [None] * 4)[:4]
                if not isinstance(item, str) or not item.strip():
                    continue
                if classification not in {"Asset", "Liability"}:
                    continue
                if amount is None:
                    continue
                key = " ".join(item.split())
                amounts[key] = Decimal(str(amount))
                classifications[key] = str(classification)
            return amounts, classifications
        finally:
            workbook.close()

    def calculate_current_snapshot(
        self,
        workbook_path: str | Path,
        annual_kpis: tuple[AnnualKPI, ...],
    ) -> CurrentSnapshot:
        amounts, classifications = self.extract_balance_sheet(workbook_path)
        total_assets = sum(
            (amount for item, amount in amounts.items() if classifications[item] == "Asset"),
            ZERO,
        )
        total_liabilities = sum(
            (
                amount
                for item, amount in amounts.items()
                if classifications[item] == "Liability"
            ),
            ZERO,
        )
        financial_assets = sum(
            (
                amount
                for item, amount in amounts.items()
                if classifications[item] == "Asset"
                and item.startswith(self.FINANCIAL_ASSET_PREFIXES)
            ),
            ZERO,
        )
        savings_cash = amounts.get("Savings", ZERO)
        house = amounts.get("House", ZERO)
        mortgage = amounts.get("Mortgage", ZERO)
        loc = amounts.get("Credit Line", ZERO)
        vehicle_loan = amounts.get("Truck loan", ZERO)

        complete = [item for item in annual_kpis if item.coverage_status == "Complete"]
        if not complete:
            raise ValueError("No complete year is available for current KPI calculations.")
        latest = max(complete, key=lambda item: item.year)
        monthly_core = latest.core_spending / Decimal("12") if latest.core_spending else ZERO
        emergency_months = (
            savings_cash / monthly_core if monthly_core > 0 else None
        )
        unsecured_debt_ratio = self._ratio(loc, latest.true_income)
        fpi_score = self._calculate_fpi(
            fixed_cost_ratio=latest.fixed_cost_ratio,
            emergency_fund_months=emergency_months,
            unsecured_debt_ratio=unsecured_debt_ratio,
            known_cash_margin=self._ratio(
                latest.true_income
                - latest.known_operating_expenses
                - latest.wealth_building
                - latest.targeted_debt_reduction,
                latest.true_income,
            ),
        )
        return CurrentSnapshot(
            source_sheet="A & L",
            latest_complete_year=latest.year,
            total_assets=total_assets,
            total_liabilities=total_liabilities,
            net_worth=total_assets - total_liabilities,
            financial_assets=financial_assets,
            savings_cash=savings_cash,
            home_equity=house - mortgage,
            mortgage_balance=mortgage,
            line_of_credit_balance=loc,
            vehicle_loan_balance=vehicle_loan,
            emergency_fund_months=emergency_months,
            unsecured_debt_ratio=unsecured_debt_ratio,
            fpi_score=fpi_score,
            fpi_band=self._fpi_band(fpi_score),
        )

    @staticmethod
    def _linear_pressure(value: Decimal, low: Decimal, high: Decimal) -> Decimal:
        if value <= low:
            return ZERO
        if value >= high:
            return ONE_HUNDRED
        return ((value - low) / (high - low)) * ONE_HUNDRED

    def _calculate_fpi(
        self,
        *,
        fixed_cost_ratio: Decimal | None,
        emergency_fund_months: Decimal | None,
        unsecured_debt_ratio: Decimal | None,
        known_cash_margin: Decimal | None,
    ) -> Decimal | None:
        if None in {
            fixed_cost_ratio,
            emergency_fund_months,
            unsecured_debt_ratio,
            known_cash_margin,
        }:
            return None
        fixed_score = self._linear_pressure(
            fixed_cost_ratio, Decimal("0.30"), Decimal("0.55")
        )
        emergency_score = ONE_HUNDRED - self._linear_pressure(
            emergency_fund_months, ZERO, Decimal("6")
        )
        debt_score = self._linear_pressure(
            unsecured_debt_ratio, ZERO, Decimal("0.15")
        )
        # A margin of 20% or more receives no pressure; zero or negative receives 100.
        cash_margin_score = ONE_HUNDRED - self._linear_pressure(
            known_cash_margin, ZERO, Decimal("0.20")
        )
        score = (
            fixed_score * Decimal("0.30")
            + emergency_score * Decimal("0.25")
            + debt_score * Decimal("0.30")
            + cash_margin_score * Decimal("0.15")
        )
        return score.quantize(Decimal("0.1"))

    @staticmethod
    def _fpi_band(score: Decimal | None) -> str:
        if score is None:
            return "Unavailable"
        if score < 25:
            return "Low"
        if score < 50:
            return "Moderate"
        if score < 75:
            return "High"
        return "Very High"
