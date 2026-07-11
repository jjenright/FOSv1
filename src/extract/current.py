"""Extractor for the current annual-budget worksheet layout."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models import Transaction
from src.transform import CategoryRegistry, normalize_category_label


SECTION_INCOME = "income"
SECTION_TRANSFERS = "transfers"
SECTION_VARIABLE_EXPENSES = "variable_expenses"
SECTION_FIXED_EXPENSES = "fixed_expenses"

_CATEGORY_TYPE_TO_SECTION = {
    "Income": SECTION_INCOME,
    "Transfer": SECTION_TRANSFERS,
    "Fixed Expense": SECTION_FIXED_EXPENSES,
    "Variable Expense": SECTION_VARIABLE_EXPENSES,
    "Irregular Expense": SECTION_VARIABLE_EXPENSES,
    "Capital": SECTION_VARIABLE_EXPENSES,
}

_CONTROL_LABELS = {
    "prev balance",
    "previous balance",
    "total balance+pay",
    "chequing ending balance",
    "next period's ending balance",
}


@dataclass(frozen=True, slots=True)
class ExtractedRecord:
    """One normalized transaction together with its reporting section."""

    section: str
    transaction: Transaction


@dataclass(frozen=True, slots=True)
class SourceRow:
    """One non-zero financial row read from the source worksheet."""

    period: str
    original_name: str
    amount: Decimal
    source_sheet: str
    source_cell: str


@dataclass(frozen=True, slots=True)
class UnknownCategory:
    """A non-zero workbook row that has no production dictionary mapping."""

    period: str
    original_name: str
    amount: Decimal
    source_sheet: str
    source_cell: str


@dataclass(frozen=True, slots=True)
class CurrentExtractionResult:
    """Result returned by :class:`CurrentLayoutExtractor`."""

    records: tuple[ExtractedRecord, ...]
    unknown_categories: tuple[UnknownCategory, ...]
    source_rows: tuple[SourceRow, ...]
    periods: tuple[str, ...]

    def for_section(self, section: str) -> tuple[Transaction, ...]:
        """Return normalized transactions belonging to one reporting section."""

        return tuple(record.transaction for record in self.records if record.section == section)

    @property
    def income(self) -> tuple[Transaction, ...]:
        return self.for_section(SECTION_INCOME)

    @property
    def transfers(self) -> tuple[Transaction, ...]:
        return self.for_section(SECTION_TRANSFERS)

    @property
    def variable_expenses(self) -> tuple[Transaction, ...]:
        return self.for_section(SECTION_VARIABLE_EXPENSES)

    @property
    def fixed_expenses(self) -> tuple[Transaction, ...]:
        return self.for_section(SECTION_FIXED_EXPENSES)


class CurrentLayoutExtractor:
    """Extract normalized records from the 2018–2026 worksheet layout."""

    _COLUMN_PAIRS = (("A", "B"), ("C", "D"), ("E", "F"))

    def __init__(self, category_registry: CategoryRegistry) -> None:
        self.category_registry = category_registry

    def extract(
        self,
        workbook_path: str | Path,
        sheet_name: str = "2025",
        *,
        include_zero: bool = False,
    ) -> CurrentExtractionResult:
        """Read one current-layout sheet from an Excel workbook."""

        path = Path(workbook_path)
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")

        workbook = load_workbook(path, data_only=True, read_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet not found: {sheet_name}")
            return self.extract_worksheet(
                workbook[sheet_name], include_zero=include_zero
            )
        finally:
            workbook.close()

    def extract_worksheet(
        self,
        worksheet: Worksheet,
        *,
        include_zero: bool = False,
    ) -> CurrentExtractionResult:
        """Extract records from an already-open current-layout worksheet."""

        year = self._sheet_year(worksheet.title)
        headers = self._period_headers(worksheet)
        if not headers:
            raise ValueError(
                f"Worksheet '{worksheet.title}' does not match the current layout."
            )

        records: list[ExtractedRecord] = []
        unknowns: list[UnknownCategory] = []
        source_rows: list[SourceRow] = []
        periods = tuple(period for _, period in headers)

        for index, (header_row, period) in enumerate(headers):
            block_end = (
                headers[index + 1][0] - 1
                if index + 1 < len(headers)
                else worksheet.max_row
            )
            block_records, block_unknowns, block_source_rows = self._extract_block(
                worksheet=worksheet,
                year=year,
                period=period,
                start_row=header_row + 1,
                end_row=block_end,
                include_zero=include_zero,
            )
            records.extend(block_records)
            unknowns.extend(block_unknowns)
            source_rows.extend(block_source_rows)

        return CurrentExtractionResult(
            records=tuple(records),
            unknown_categories=tuple(unknowns),
            source_rows=tuple(source_rows),
            periods=periods,
        )

    def _extract_block(
        self,
        *,
        worksheet: Worksheet,
        year: int,
        period: str,
        start_row: int,
        end_row: int,
        include_zero: bool,
    ) -> tuple[list[ExtractedRecord], list[UnknownCategory], list[SourceRow]]:
        records: list[ExtractedRecord] = []
        unknowns: list[UnknownCategory] = []
        source_rows: list[SourceRow] = []

        for label_column, amount_column in self._COLUMN_PAIRS:
            for row_number in range(start_row, end_row + 1):
                label_cell = worksheet[f"{label_column}{row_number}"]
                amount_cell = worksheet[f"{amount_column}{row_number}"]
                label = label_cell.value
                amount_value = amount_cell.value

                if not isinstance(label, str) or not self._is_number(amount_value):
                    continue

                amount = Decimal(str(amount_value))
                if amount == 0 and not include_zero:
                    continue
                if self._is_control_label(label):
                    continue

                cleaned_label = " ".join(label.split())
                source_rows.append(
                    SourceRow(
                        period=period,
                        original_name=cleaned_label,
                        amount=amount,
                        source_sheet=worksheet.title,
                        source_cell=label_cell.coordinate,
                    )
                )

                category = self.category_registry.find(label)
                if category is None:
                    unknowns.append(
                        UnknownCategory(
                            period=period,
                            original_name=cleaned_label,
                            amount=amount,
                            source_sheet=worksheet.title,
                            source_cell=label_cell.coordinate,
                        )
                    )
                    continue

                section = _CATEGORY_TYPE_TO_SECTION.get(category.category_type)
                if section is None:
                    raise ValueError(
                        f"Unsupported category type '{category.category_type}' "
                        f"for '{label}'."
                    )

                transaction = Transaction(
                    year=year,
                    period=period,
                    category_id=category.category_id,
                    description=category.display_name,
                    amount=amount,
                    source_sheet=worksheet.title,
                    source_cell=label_cell.coordinate,
                )
                records.append(ExtractedRecord(section=section, transaction=transaction))

        return records, unknowns, source_rows

    @staticmethod
    def _period_headers(worksheet: Worksheet) -> list[tuple[int, str]]:
        headers: list[tuple[int, str]] = []
        for row_number in range(1, worksheet.max_row + 1):
            period = worksheet[f"A{row_number}"].value
            if (
                isinstance(period, str)
                and worksheet[f"B{row_number}"].value == "PayAmount"
                and worksheet[f"C{row_number}"].value == "Bi-weekly"
                and worksheet[f"E{row_number}"].value == "Monthlies"
            ):
                headers.append((row_number, " ".join(period.split())))
        return headers

    @staticmethod
    def _sheet_year(sheet_name: str) -> int:
        try:
            return int(sheet_name[:4])
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Worksheet name does not begin with a year: {sheet_name}") from exc

    @staticmethod
    def _is_number(value: object) -> bool:
        return isinstance(value, (int, float)) and not isinstance(value, bool)

    @staticmethod
    def _is_control_label(label: str) -> bool:
        return normalize_category_label(label) in _CONTROL_LABELS
