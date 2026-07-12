"""Extractor for detailed Visa transaction sheets copied into the source workbook."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from src.extract.current import (
    SECTION_FIXED_EXPENSES,
    SECTION_INCOME,
    SECTION_TRANSFERS,
    SECTION_VARIABLE_EXPENSES,
    CurrentExtractionResult,
    ExtractedRecord,
    SourceRow,
    UnknownCategory,
)
from src.models import Transaction
from src.transform import CategoryRegistry

_HEADER_REQUIRED = {
    "transactionid",
    "transactiondate",
    "amount",
    "transactiontype",
}

_CATEGORY_TYPE_TO_SECTION = {
    "Income": SECTION_INCOME,
    "Transfer": SECTION_TRANSFERS,
    "Fixed Expense": SECTION_FIXED_EXPENSES,
    "Variable Expense": SECTION_VARIABLE_EXPENSES,
    "Irregular Expense": SECTION_VARIABLE_EXPENSES,
    "Capital": SECTION_VARIABLE_EXPENSES,
}


@dataclass(frozen=True, slots=True)
class VisaImportStats:
    """Audit metrics for transaction-detail sheets merged into annual history."""

    source_sheets: tuple[str, ...] = ()
    imported_records: int = 0
    imported_total: Decimal = Decimal("0")
    skipped_transfers: int = 0
    duplicate_transaction_ids: int = 0
    unknown_records: int = 0


@dataclass(frozen=True, slots=True)
class VisaExtractionResult:
    """Detailed transaction rows grouped by transaction year."""

    by_year: dict[int, CurrentExtractionResult]
    stats: VisaImportStats


class VisaSheetExtractor:
    """Read a copied ``Import_2025``-style sheet using its column headers."""

    def __init__(self, category_registry: CategoryRegistry) -> None:
        self.registry = category_registry
        self.entries = {
            str(entry["category_id"]): entry for entry in category_registry.entries()
        }

    @staticmethod
    def _normalize_header(value: Any) -> str:
        if value is None:
            return ""
        return re.sub(r"[^a-z0-9]+", "", str(value).casefold())

    @classmethod
    def header_map(cls, worksheet: Worksheet) -> tuple[int, dict[str, int]] | None:
        """Return the header row and normalized column map when the sheet matches."""

        for row_number in range(1, min(worksheet.max_row, 20) + 1):
            mapping: dict[str, int] = {}
            for cell in worksheet[row_number]:
                normalized = cls._normalize_header(cell.value)
                if normalized:
                    mapping[normalized] = cell.column
            if _HEADER_REQUIRED.issubset(mapping) and (
                "categoryid" in mapping or "suggestedcategory" in mapping
            ):
                return row_number, mapping
        return None

    @classmethod
    def matches(cls, worksheet: Worksheet) -> bool:
        return cls.header_map(worksheet) is not None

    def extract_workbook_sheets(self, workbook: Any) -> VisaExtractionResult:
        grouped_records: dict[int, list[ExtractedRecord]] = {}
        grouped_unknowns: dict[int, list[UnknownCategory]] = {}
        grouped_sources: dict[int, list[SourceRow]] = {}
        grouped_periods: dict[int, set[str]] = {}
        source_sheets: list[str] = []
        seen_ids: set[str] = set()
        duplicate_ids = 0
        skipped_transfers = 0
        imported_total = Decimal("0")

        for worksheet in workbook.worksheets:
            header = self.header_map(worksheet)
            if header is None:
                continue
            source_sheets.append(worksheet.title)
            header_row, columns = header

            for row_number in range(header_row + 1, worksheet.max_row + 1):
                transaction_id = self._cell_text(worksheet, row_number, columns, "transactionid")
                transaction_date = self._parse_date(
                    self._cell_value(worksheet, row_number, columns, "transactiondate")
                )
                amount = self._parse_decimal(
                    self._cell_value(worksheet, row_number, columns, "amount")
                )
                transaction_type = self._cell_text(
                    worksheet, row_number, columns, "transactiontype"
                )

                if not transaction_id and transaction_date is None and amount is None:
                    continue
                if transaction_date is None or amount is None:
                    raise ValueError(
                        f"Visa transaction row {worksheet.title}!{row_number} is missing "
                        "a valid transaction date or amount."
                    )

                if transaction_id:
                    if transaction_id in seen_ids:
                        duplicate_ids += 1
                        continue
                    seen_ids.add(transaction_id)

                if transaction_type.casefold() == "transfer":
                    skipped_transfers += 1
                    continue

                year = transaction_date.year
                period = f"Visa {transaction_date:%Y-%m}"
                activity = self._cell_text(
                    worksheet, row_number, columns, "activitydescription"
                )
                merchant = self._cell_text(
                    worksheet, row_number, columns, "normalizedmerchant"
                )
                raw_label = merchant or activity or "Visa transaction"
                source_col = columns.get("transactionid", 1)
                source_cell = worksheet.cell(row_number, source_col).coordinate
                category_id = self._cell_text(
                    worksheet, row_number, columns, "categoryid"
                )
                suggested = self._cell_text(
                    worksheet, row_number, columns, "suggestedcategory"
                )

                grouped_sources.setdefault(year, []).append(
                    SourceRow(
                        period=period,
                        original_name=raw_label,
                        amount=amount,
                        source_sheet=worksheet.title,
                        source_cell=source_cell,
                    )
                )
                grouped_periods.setdefault(year, set()).add(period)
                imported_total += amount

                entry = self.entries.get(category_id)
                if entry is None and suggested:
                    category = self.registry.find(suggested)
                    if category is not None:
                        category_id = category.category_id
                        entry = self.entries.get(category_id)

                if entry is None:
                    grouped_unknowns.setdefault(year, []).append(
                        UnknownCategory(
                            period=period,
                            original_name=raw_label,
                            amount=amount,
                            source_sheet=worksheet.title,
                            source_cell=source_cell,
                        )
                    )
                    continue

                section = _CATEGORY_TYPE_TO_SECTION.get(str(entry["category_type"]))
                if section is None:
                    raise ValueError(
                        f"Unsupported Visa category type '{entry['category_type']}' "
                        f"at {worksheet.title}!{source_cell}."
                    )
                if section == SECTION_TRANSFERS:
                    skipped_transfers += 1
                    grouped_sources[year].pop()
                    imported_total -= amount
                    continue

                description = raw_label or str(entry["display_name"])
                transaction = Transaction(
                    year=year,
                    period=period,
                    category_id=category_id,
                    description=description,
                    amount=amount,
                    source_sheet=worksheet.title,
                    source_cell=source_cell,
                    transaction_date=transaction_date,
                )
                grouped_records.setdefault(year, []).append(
                    ExtractedRecord(section=section, transaction=transaction)
                )

        by_year: dict[int, CurrentExtractionResult] = {}
        years = sorted(set(grouped_sources) | set(grouped_records) | set(grouped_unknowns))
        for year in years:
            by_year[year] = CurrentExtractionResult(
                records=tuple(grouped_records.get(year, [])),
                unknown_categories=tuple(grouped_unknowns.get(year, [])),
                source_rows=tuple(grouped_sources.get(year, [])),
                periods=tuple(sorted(grouped_periods.get(year, set()))),
            )

        imported_records = sum(len(result.records) for result in by_year.values())
        unknown_records = sum(len(result.unknown_categories) for result in by_year.values())
        return VisaExtractionResult(
            by_year=by_year,
            stats=VisaImportStats(
                source_sheets=tuple(source_sheets),
                imported_records=imported_records,
                imported_total=imported_total,
                skipped_transfers=skipped_transfers,
                duplicate_transaction_ids=duplicate_ids,
                unknown_records=unknown_records,
            ),
        )

    @staticmethod
    def _cell_value(
        worksheet: Worksheet, row: int, columns: dict[str, int], name: str
    ) -> Any:
        column = columns.get(name)
        return None if column is None else worksheet.cell(row, column).value

    @classmethod
    def _cell_text(
        cls, worksheet: Worksheet, row: int, columns: dict[str, int], name: str
    ) -> str:
        value = cls._cell_value(worksheet, row, columns, name)
        return "" if value is None else " ".join(str(value).split())

    @staticmethod
    def _parse_date(value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            cleaned = value.strip()
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(cleaned, fmt).date()
                except ValueError:
                    continue
        return None

    @staticmethod
    def _parse_decimal(value: Any) -> Decimal | None:
        if value is None or isinstance(value, bool):
            return None
        try:
            if isinstance(value, str):
                cleaned = value.replace("$", "").replace(",", "").strip()
                if cleaned.startswith("(") and cleaned.endswith(")"):
                    cleaned = "-" + cleaned[1:-1]
                return Decimal(cleaned)
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None
