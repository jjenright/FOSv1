"""Workbook-wide historical extraction orchestration."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.extract.current import CurrentExtractionResult, CurrentLayoutExtractor
from src.extract.layout_detector import LayoutDetector
from src.extract.legacy import LegacyLayoutExtractor
from src.extract.visa import VisaImportStats, VisaSheetExtractor
from src.transform import CategoryRegistry


@dataclass(frozen=True, slots=True)
class SheetExtraction:
    """Extraction result and metadata for one annual worksheet."""

    sheet_name: str
    layout: str
    result: CurrentExtractionResult

    @property
    def year(self) -> int:
        return int(self.sheet_name[:4])


@dataclass(frozen=True, slots=True)
class HistoricalExtractionResult:
    """Normalized extraction results for all official annual worksheets."""

    sheets: tuple[SheetExtraction, ...]
    excluded_sheets: tuple[str, ...] = ()
    visa_import: VisaImportStats | None = None

    @property
    def records(self):
        return tuple(record for sheet in self.sheets for record in sheet.result.records)

    @property
    def income(self):
        return tuple(item for sheet in self.sheets for item in sheet.result.income)

    @property
    def transfers(self):
        return tuple(item for sheet in self.sheets for item in sheet.result.transfers)

    @property
    def variable_expenses(self):
        return tuple(
            item for sheet in self.sheets for item in sheet.result.variable_expenses
        )

    @property
    def fixed_expenses(self):
        return tuple(item for sheet in self.sheets for item in sheet.result.fixed_expenses)

    @property
    def unknown_categories(self):
        return tuple(
            item for sheet in self.sheets for item in sheet.result.unknown_categories
        )

    @property
    def source_rows(self):
        return tuple(item for sheet in self.sheets for item in sheet.result.source_rows)

    @property
    def periods(self):
        return tuple(
            f"{sheet.sheet_name}:{period}"
            for sheet in self.sheets
            for period in sheet.result.periods
        )


class HistoricalWorkbookExtractor:
    """Extract every official annual worksheet without double-counting archives."""

    def __init__(
        self,
        layout_detector: LayoutDetector,
        category_registry: CategoryRegistry,
    ) -> None:
        self.layout_detector = layout_detector
        self.category_registry = category_registry
        self.current = CurrentLayoutExtractor(category_registry)
        self.legacy = LegacyLayoutExtractor(category_registry)
        self.visa = VisaSheetExtractor(category_registry)

    def official_sheets(self) -> tuple[str, ...]:
        sheets = [
            sheet
            for sheet in self.layout_detector.configured_sheets()
            if not sheet.casefold().endswith("(old)")
        ]
        return tuple(sorted(sheets, key=lambda item: (int(item[:4]), item)))

    def extract(
        self,
        workbook_path: str | Path,
        *,
        sheets: tuple[str, ...] | list[str] | None = None,
        include_zero: bool = False,
    ) -> HistoricalExtractionResult:
        source = Path(workbook_path)
        if not source.is_file():
            raise FileNotFoundError(f"Workbook not found: {source}")

        selected = tuple(sheets) if sheets is not None else self.official_sheets()
        excluded = tuple(
            sheet
            for sheet in self.layout_detector.configured_sheets()
            if sheet not in selected
        )

        workbook = load_workbook(source, data_only=True, read_only=False)
        try:
            missing = [sheet for sheet in selected if sheet not in workbook.sheetnames]
            if missing:
                raise ValueError(
                    "Configured annual worksheets are missing: " + ", ".join(missing)
                )

            extracted: list[SheetExtraction] = []
            for sheet_name in selected:
                worksheet = workbook[sheet_name]
                configured_layout = self.layout_detector.detect(sheet_name)
                if CurrentLayoutExtractor._period_headers(worksheet):
                    result = self.current.extract_worksheet(
                        worksheet, include_zero=include_zero
                    )
                    effective_layout = "current"
                elif LegacyLayoutExtractor._period_headers(worksheet):
                    result = self.legacy.extract_worksheet(
                        worksheet, include_zero=include_zero
                    )
                    effective_layout = "legacy"
                else:
                    raise ValueError(
                        f"Worksheet '{sheet_name}' does not match a supported layout."
                    )

                # Preserve configured era metadata while documenting the actual
                # extractor used for mixed transitional sheets.
                layout = (
                    configured_layout
                    if configured_layout == effective_layout
                    else f"{configured_layout}/{effective_layout}"
                )
                extracted.append(
                    SheetExtraction(
                        sheet_name=sheet_name,
                        layout=layout,
                        result=result,
                    )
                )

            visa_extraction = self.visa.extract_workbook_sheets(workbook)
            if visa_extraction.by_year:
                by_year = {item.year: index for index, item in enumerate(extracted)}
                missing_years = sorted(set(visa_extraction.by_year) - set(by_year))
                if missing_years:
                    raise ValueError(
                        "Detailed Visa transactions were found for years without an "
                        "imported annual worksheet: " + ", ".join(str(year) for year in missing_years)
                    )
                for year, visa_result in visa_extraction.by_year.items():
                    index = by_year[year]
                    annual = extracted[index]
                    merged = CurrentExtractionResult(
                        records=annual.result.records + visa_result.records,
                        unknown_categories=(
                            annual.result.unknown_categories + visa_result.unknown_categories
                        ),
                        source_rows=annual.result.source_rows + visa_result.source_rows,
                        # Visa months are not pay periods and must not alter annual coverage.
                        periods=annual.result.periods,
                    )
                    extracted[index] = SheetExtraction(
                        sheet_name=annual.sheet_name,
                        layout=f"{annual.layout}+visa",
                        result=merged,
                    )
            visa_stats = visa_extraction.stats
        finally:
            workbook.close()

        return HistoricalExtractionResult(
            sheets=tuple(extracted),
            excluded_sheets=excluded,
            visa_import=visa_stats,
        )
