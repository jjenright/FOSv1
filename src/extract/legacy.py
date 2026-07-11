"""Extractor for the legacy annual-budget worksheet layout."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.extract.current import (
    SECTION_FIXED_EXPENSES,
    SECTION_INCOME,
    SECTION_TRANSFERS,
    SECTION_VARIABLE_EXPENSES,
    CurrentExtractionResult,
    ExtractedRecord,
    SourceRow,
    UnknownCategory,
    _CATEGORY_TYPE_TO_SECTION,
)
from src.models import Transaction
from src.transform import CategoryRegistry, normalize_category_label

_CONTROL_ROW_LABELS = {
    normalize_category_label(label)
    for label in (
        "Pay Period",
        "Chequing",
        "Chequing Beginning Balance",
        "Chequing Ending Balance",
        "Previous Balance",
        "Prev balance",
        "Starting Balance",
        "Savings Beginning Balance",
        "Total",
        "Total +",
        "Total -",
        "Total Balance+Pay",
        "Total Spent During Period",
        "Bills During Period",
        "Total Ending Balance",
        "Next Period's Ending Balance",
        "Visa",
        "Visa Opening Balance",
        "Visa Beginning Balance",
        "Visa Ending Balance",
        "Visa closing balance",
        "Mastercard",
        "Mastercard Opening Balance",
        "Mastercard Beginning Balance",
        "Mastercard Ending Balance",
        "Alotted Items Deficit",
        "Allotted Items Deficit",
        "Bi-Weekly Deductions",
        "Monthly Deductions",
        "One-Time Deductions",
        "Purchases",
        "Payments",
    )
}

_CONTROL_ANNOTATIONS = {
    normalize_category_label(label)
    for label in (
        "Not including Alotted Items Deficit",
        "Not including Allotted Items Deficit",
        "$800 target",
        "Auto",
        "Used",
        "Remaining allotted",
    )
}


class LegacyLayoutExtractor:
    """Extract normalized records from the 2008–2016 legacy layout."""

    def __init__(self, category_registry: CategoryRegistry) -> None:
        self.category_registry = category_registry

    def extract(
        self,
        workbook_path: str | Path,
        sheet_name: str,
        *,
        include_zero: bool = False,
    ) -> CurrentExtractionResult:
        path = Path(workbook_path)
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")

        workbook = load_workbook(path, data_only=True, read_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet not found: {sheet_name}")
            return self.extract_worksheet(
                workbook[sheet_name], include_zero=include_zero
            )
        finally:
            workbook.close()

    def extract_worksheet(
        self,
        worksheet: Worksheet,
        *,
        include_zero: bool = False,
    ) -> CurrentExtractionResult:
        year = self._sheet_year(worksheet.title)
        headers = self._period_headers(worksheet)
        if not headers:
            raise ValueError(
                f"Worksheet '{worksheet.title}' does not match the legacy layout."
            )

        records: list[ExtractedRecord] = []
        unknowns: list[UnknownCategory] = []
        source_rows: list[SourceRow] = []

        for index, (header_row, period) in enumerate(headers):
            block_end = (
                headers[index + 1][0] - 1
                if index + 1 < len(headers)
                else min(worksheet.max_row, header_row + 120)
            )
            block_records, block_unknowns, block_source_rows = self._extract_block(
                worksheet=worksheet,
                year=year,
                period=period,
                start_row=header_row + 1,
                end_row=block_end,
                include_zero=include_zero,
            )
            records.extend(block_records)
            unknowns.extend(block_unknowns)
            source_rows.extend(block_source_rows)

        return CurrentExtractionResult(
            records=tuple(records),
            unknown_categories=tuple(unknowns),
            source_rows=tuple(source_rows),
            periods=tuple(period for _, period in headers),
        )

    def _extract_block(
        self,
        *,
        worksheet: Worksheet,
        year: int,
        period: str,
        start_row: int,
        end_row: int,
        include_zero: bool,
    ) -> tuple[list[ExtractedRecord], list[UnknownCategory], list[SourceRow]]:
        records: list[ExtractedRecord] = []
        unknowns: list[UnknownCategory] = []
        source_rows: list[SourceRow] = []

        first_deductions_row = self._first_section_row(worksheet, start_row, end_row)

        for row_number in range(start_row, end_row + 1):
            a_value = worksheet.cell(row_number, 1).value
            b_value = worksheet.cell(row_number, 2).value
            c_value = worksheet.cell(row_number, 3).value

            if not self._is_number(b_value):
                continue
            amount = Decimal(str(b_value))
            if amount == 0 and not include_zero:
                continue

            a_label = self._clean_label(a_value)
            c_label = self._clean_label(c_value)
            if a_label and normalize_category_label(a_label) in _CONTROL_ROW_LABELS:
                continue

            # In legacy sheets, the category normally appears in column C and
            # the amount in column B. This includes credit-card purchase detail.
            if c_label:
                if normalize_category_label(c_label) in _CONTROL_ANNOTATIONS:
                    continue
                self._classify(
                    worksheet=worksheet,
                    year=year,
                    period=period,
                    label=c_label,
                    amount=amount,
                    source_cell=f"C{row_number}",
                    records=records,
                    unknowns=unknowns,
                    source_rows=source_rows,
                )
                continue

            # Column A contains income and adjustments before the first
            # deduction section. Mapped labels elsewhere are also retained.
            if not a_label:
                continue
            category = self.category_registry.find(a_label)
            if category is None and row_number >= first_deductions_row:
                continue
            if any(
                token in normalize_category_label(a_label)
                for token in ("balance", "opening", "ending", "total")
            ) and category is None:
                continue

            self._classify(
                worksheet=worksheet,
                year=year,
                period=period,
                label=a_label,
                amount=amount,
                source_cell=f"A{row_number}",
                records=records,
                unknowns=unknowns,
                source_rows=source_rows,
            )

        return records, unknowns, source_rows

    def _classify(
        self,
        *,
        worksheet: Worksheet,
        year: int,
        period: str,
        label: str,
        amount: Decimal,
        source_cell: str,
        records: list[ExtractedRecord],
        unknowns: list[UnknownCategory],
        source_rows: list[SourceRow],
    ) -> None:
        source_rows.append(
            SourceRow(
                period=period,
                original_name=label,
                amount=amount,
                source_sheet=worksheet.title,
                source_cell=source_cell,
            )
        )
        category = self.category_registry.find(label)
        if category is None:
            unknowns.append(
                UnknownCategory(
                    period=period,
                    original_name=label,
                    amount=amount,
                    source_sheet=worksheet.title,
                    source_cell=source_cell,
                )
            )
            return

        section = _CATEGORY_TYPE_TO_SECTION.get(category.category_type)
        if section is None:
            raise ValueError(
                f"Unsupported category type '{category.category_type}' for '{label}'."
            )
        records.append(
            ExtractedRecord(
                section=section,
                transaction=Transaction(
                    year=year,
                    period=period,
                    category_id=category.category_id,
                    description=category.display_name,
                    amount=amount,
                    source_sheet=worksheet.title,
                    source_cell=source_cell,
                ),
            )
        )

    @staticmethod
    def _first_section_row(worksheet: Worksheet, start_row: int, end_row: int) -> int:
        for row_number in range(start_row, end_row + 1):
            value = worksheet.cell(row_number, 1).value
            if isinstance(value, str) and normalize_category_label(value) in {
                normalize_category_label("Bi-Weekly Deductions"),
                normalize_category_label("Monthly Deductions"),
            }:
                return row_number
        return end_row + 1

    @staticmethod
    def _period_headers(worksheet: Worksheet) -> list[tuple[int, str]]:
        headers: list[tuple[int, str]] = []
        counts: dict[str, int] = {}
        max_scan_row = min(worksheet.max_row, 5000)
        for row_number in range(1, max_scan_row + 1):
            if worksheet.cell(row_number, 1).value != "Pay Period":
                continue
            period = worksheet.cell(row_number, 2).value
            if isinstance(period, str):
                cleaned = " ".join(period.split())
                counts[cleaned] = counts.get(cleaned, 0) + 1
                unique = (
                    cleaned
                    if counts[cleaned] == 1
                    else f"{cleaned} [{counts[cleaned]}]"
                )
                headers.append((row_number, unique))
        return headers

    @staticmethod
    def _sheet_year(sheet_name: str) -> int:
        try:
            return int(sheet_name[:4])
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Worksheet name does not begin with a year: {sheet_name}"
            ) from exc

    @staticmethod
    def _clean_label(value: object) -> str | None:
        if not isinstance(value, str):
            return None
        cleaned = " ".join(value.split())
        return cleaned or None

    @staticmethod
    def _is_number(value: object) -> bool:
        return isinstance(value, (int, float)) and not isinstance(value, bool)
