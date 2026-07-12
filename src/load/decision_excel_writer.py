"""Excel sheets for FOS v1.1.0 decision intelligence."""
from __future__ import annotations
from decimal import Decimal
from typing import Any
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from src.decision import DecisionReport
from src.extract import HistoricalExtractionResult
from src.load.excel_loader import BLACK,CURRENCY_FORMAT,DARK_BLUE,INTEGER_FORMAT,LIGHT_BLUE,LIGHT_GREEN,LIGHT_RED,LIGHT_YELLOW,WHITE
from src.transform import CategoryRegistry
PCT='0.0%;[Red](0.0%);-'
class DecisionExcelWriter:
    def __init__(self,registry:CategoryRegistry)->None:
        self.registry=registry; self.entries={str(e['category_id']):e for e in registry.entries()}
    def write_all(self,wb:Any,r:DecisionReport,x:HistoricalExtractionResult)->None:
        self._decision(wb['Decision_Centre'],r); self._opps(wb['Opportunities'],r); self._debt(wb['Debt_Planner'],r); self._forecast(wb['Forecast_12M'],r); self._dna(wb['Financial_DNA'],r); self._history(wb['Category_History'],r); self._explorer(wb['Spending_Explorer'],r,x); self._merchants(wb['Merchant_Intelligence'],r); self._dashboard(wb['Dashboard'],r)
    @staticmethod
    def _title(ws:Any,title:str,subtitle:str)->None:
        ws.sheet_view.showGridLines=False; ws.merge_cells('A1:L2'); ws['A1']=title; ws['A1'].font=Font(size=18,bold=True,color=WHITE); ws['A1'].fill=PatternFill('solid',fgColor=DARK_BLUE); ws['A1'].alignment=Alignment(vertical='center'); ws.merge_cells('A3:L3'); ws['A3']=subtitle; ws['A3'].font=Font(italic=True,color='44546A')
    @staticmethod
    def _section(ws:Any,row:int,title:str,end:int=12)->None:
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=end); c=ws.cell(row,1,title); c.font=Font(bold=True,color=WHITE); c.fill=PatternFill('solid',fgColor=DARK_BLUE)
    @staticmethod
    def _table(ws:Any,row:int,col:int,headers:list[str],rows:list[list[Any]],name:str)->Table:
        for i,h in enumerate(headers):
            c=ws.cell(row,col+i,h); c.font=Font(bold=True,color=WHITE); c.fill=PatternFill('solid',fgColor=DARK_BLUE); c.alignment=Alignment(horizontal='center',vertical='center')
        data=rows or [[None]*len(headers)]
        for ri,values in enumerate(data,1):
            for ci,v in enumerate(values): ws.cell(row+ri,col+ci,v)
        endr=row+len(data); endc=col+len(headers)-1; ref=f'{get_column_letter(col)}{row}:{get_column_letter(endc)}{endr}'; t=Table(displayName=name,ref=ref); t.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False); ws.add_table(t); return t
    @staticmethod
    def _widths(ws:Any,widths:list[float],start:int=1)->None:
        for i,w in enumerate(widths,start): ws.column_dimensions[get_column_letter(i)].width=w
    def _decision(self,ws,r):
        self._title(ws,'Decision Centre','Direct spending choices, LOC acceleration, forecast outcomes and long-term financial patterns.'); ws.freeze_panes='A6'; self._section(ws,5,'Highest-impact spending opportunities')
        rows=[[o.opportunity,float(o.annual_spend),float(o.save_50),float(o.save_100),o.loc_weeks_saved_50,o.direct_insight] for o in r.opportunities[:5]]; self._table(ws,6,1,['Opportunity','Annual Spend ($)','Save 50% ($)','Save 100% ($)','LOC Weeks Saved at 50%','Direct Insight'],rows,'DecisionTopOpportunitiesTable')
        for rr in range(7,7+len(rows)):
            for c in (2,3,4): ws.cell(rr,c).number_format=CURRENCY_FORMAT
            ws.cell(rr,6).alignment=Alignment(wrap_text=True,vertical='top')
        nr=8+max(1,len(rows)); self._section(ws,nr,'Scenario outcomes'); sums=[[s.scenario,float(s.annual_discretionary_savings),s.loc_payoff_month or 'Beyond horizon',float(s.ending_loc),float(s.ending_savings),float(s.annual_unallocated_margin)] for s in r.forecast_summaries]; self._table(ws,nr+1,1,['Scenario','Annual Spending Freed ($)','LOC Payoff Month','Ending LOC ($)','Ending Liquid Savings ($)','Unallocated Cash Margin ($)'],sums,'DecisionScenarioTable')
        for rr in range(nr+2,nr+2+len(sums)):
            for c in (2,4,5,6): ws.cell(rr,c).number_format=CURRENCY_FORMAT
        dr=nr+3+max(1,len(sums)); self._section(ws,dr,'Financial DNA findings'); drows=[[d.rank,d.theme,d.finding,d.evidence,d.implication] for d in r.dna_findings]; self._table(ws,dr+1,1,['Rank','Theme','Finding','Evidence','Implication'],drows,'DecisionDNAFindingsTable')
        for rr in range(dr+2,dr+2+len(drows)):
            ws.row_dimensions[rr].height=54; ws.cell(rr,1).alignment=Alignment(horizontal='center',vertical='center')
            for c in range(2,6): ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical='top')
        lr=dr+3+max(1,len(drows)); self._section(ws,lr,'Open detailed analysis'); links=[('Opportunities',"#'Opportunities'!A1"),('Debt Planner',"#'Debt_Planner'!A1"),('12-Month Forecast',"#'Forecast_12M'!A1"),('Financial DNA',"#'Financial_DNA'!A1"),('Spending Explorer',"#'Spending_Explorer'!A1"),('Merchant Intelligence',"#'Merchant_Intelligence'!A1")]
        for i,(label,target) in enumerate(links,1): c=ws.cell(lr+1,i*2-1,label); c.hyperlink=target; c.style='Hyperlink'; c.font=Font(bold=True,color='0563C1')
        self._widths(ws,[18,18,18,18,20,58,4,18,4,18,4,18])
    def _opps(self,ws,r):
        self._title(ws,'Spending Opportunity Finder',f'Latest complete-year ({r.latest_year}) discretionary spending with 25%, 50% and full-reduction scenarios.')
        rows=[[o.rank,o.category_id,o.opportunity,o.essentiality,float(o.annual_spend),float(o.monthly_spend),float(o.save_25),float(o.save_50),float(o.save_100),o.loc_weeks_saved_50,o.loc_weeks_saved_100,o.recommendation,o.direct_insight] for o in r.opportunities]; self._table(ws,5,1,['Rank','CategoryID','Opportunity','Essentiality','Annual Spend ($)','Monthly Spend ($)','Save 25% ($/yr)','Save 50% ($/yr)','Save 100% ($/yr)','LOC Weeks Saved 50%','LOC Weeks Saved 100%','Recommended Approach','Direct Insight'],rows,'SpendingOpportunitiesTable'); ws.freeze_panes='A6'
        for rr in range(6,6+len(rows)):
            for c in range(5,10): ws.cell(rr,c).number_format=CURRENCY_FORMAT
            for c in (12,13): ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical='top')
        self._widths(ws,[8,13,28,24,18,18,18,18,19,20,21,46,72])
    def _debt(self,ws,r):
        self._title(ws,'LOC Debt Payoff Planner','Scenario comparison using editable assumptions. The 0.99% vehicle loan remains outside the acceleration target.'); rate=Decimal(str(r.assumptions['loc_annual_interest_rate'])); weekly=Decimal(str(r.assumptions['loc_weekly_payment'])); self._table(ws,5,1,['Assumption','Value','Source / Note'],[['LOC annual interest rate',float(rate),'User-provided 6–7% range midpoint'],['Current weekly payment',float(weekly),'User-provided planned payment'],['Payment frequency',52,'Weekly']],'DebtAssumptionsTable'); ws['B6'].number_format=PCT; ws['B7'].number_format=CURRENCY_FORMAT
        for c in ('B6','B7'): ws[c].font=Font(color='0000FF'); ws[c].fill=PatternFill('solid',fgColor='FFF2CC')
        rows=[[d.scenario,float(d.weekly_payment),float(d.annual_redirect),d.payoff_weeks,float(d.payoff_months),float(d.interest_cost),float(d.interest_saved),d.weeks_saved,float(d.ending_balance_12m)] for d in r.debt_scenarios]; self._table(ws,10,1,['Scenario','Weekly Payment ($)','Annual Redirect ($)','Payoff Weeks','Payoff Months','Interest Cost ($)','Interest Saved ($)','Weeks Saved','Ending LOC after 12M ($)'],rows,'DebtScenarioTable')
        for rr in range(11,11+len(rows)):
            for c in (2,3,6,7,9): ws.cell(rr,c).number_format=CURRENCY_FORMAT
            ws.cell(rr,5).number_format='0.0'
        ws.freeze_panes='A11'; self._widths(ws,[38,20,20,16,16,18,19,15,24])
    def _forecast(self,ws,r):
        self._title(ws,'12-Month Cash-Flow Forecast','Latest complete-year averages; unallocated margin is shown separately and is not assumed to become savings.')
        sums=[[s.scenario,float(s.reduction_rate),float(s.annual_discretionary_savings),s.loc_payoff_month or 'Beyond horizon',float(s.ending_loc),float(s.ending_savings),float(s.annual_unallocated_margin)] for s in r.forecast_summaries]; self._table(ws,5,1,['Scenario','Reduction Rate','Annual Spending Freed ($)','LOC Payoff Month','Ending LOC ($)','Ending Savings ($)','Unallocated Margin ($)'],sums,'ForecastSummaryTable')
        for rr in range(6,6+len(sums)):
            ws.cell(rr,2).number_format=PCT
            for c in (3,5,6,7): ws.cell(rr,c).number_format=CURRENCY_FORMAT
        start=8+max(1,len(sums)); rows=[[m.scenario,m.month_number,float(m.income),float(m.operating_expenses),float(m.discretionary_reduction),float(m.wealth_building),float(m.loc_payment),float(m.loc_interest),float(m.unallocated_margin),float(m.ending_loc),float(m.ending_savings)] for m in r.forecast_months]; self._table(ws,start,1,['Scenario','Month','Income ($)','Operating Expenses ($)','Spending Reduction ($)','Wealth Building ($)','LOC Payment ($)','LOC Interest ($)','Unallocated Margin ($)','Ending LOC ($)','Ending Savings ($)'],rows,'ForecastMonthlyTable')
        for rr in range(start+1,start+1+len(rows)):
            for c in range(3,12): ws.cell(rr,c).number_format=CURRENCY_FORMAT
        ws.freeze_panes=f'A{start+1}'; self._widths(ws,[28,10,16,22,21,20,18,18,22,18,20])
        names=[s.scenario for s in r.forecast_summaries]; ws.cell(5,13,'Month')
        for i,n in enumerate(names,14): ws.cell(5,i,n)
        by={(m.scenario,m.month_number):m for m in r.forecast_months}; months=int(r.assumptions['forecast_months'])
        for m in range(1,months+1):
            ws.cell(5+m,13,m)
            for i,n in enumerate(names,14): ws.cell(5+m,i,float(by[(n,m)].ending_loc)); ws.cell(5+m,i).number_format=CURRENCY_FORMAT
        chart=LineChart(); chart.visible_cells_only=True; chart.title='LOC Balance by Forecast Scenario'; chart.y_axis.title='LOC Balance ($)'; chart.x_axis.title='Month'; chart.height=8; chart.width=14; chart.add_data(Reference(ws,min_col=14,max_col=13+len(names),min_row=5,max_row=5+months),titles_from_data=True); chart.set_categories(Reference(ws,min_col=13,min_row=6,max_row=5+months)); chart.legend.position='b'; ws.add_chart(chart,'M19')
    def _dna(self,ws,r):
        self._title(ws,'Financial DNA','Long-term household patterns based only on comparison-eligible historical years.'); finds=[[d.rank,d.theme,d.finding,d.evidence,d.implication] for d in r.dna_findings]; self._table(ws,5,1,['Rank','Theme','Finding','Evidence','Implication'],finds,'DNAFindingsTable')
        for rr in range(6,6+len(finds)):
            for c in range(2,6): ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical='top')
        start=7+max(1,len(finds)); rows=[[d.category,d.pattern,d.years_observed,d.eligible_years,float(d.persistence_ratio),d.first_year,d.latest_year,float(d.average_spend),float(d.median_spend),float(d.latest_spend),float(d.latest_share_of_income) if d.latest_share_of_income is not None else None,float(d.annual_trend),float(d.volatility) if d.volatility is not None else None,d.peak_year,float(d.peak_spend)] for d in r.category_dna]; self._table(ws,start,1,['Category','Pattern','Years Observed','Eligible Years','Persistence','First Year','Latest Year','Average Spend ($)','Median Spend ($)','Latest Spend ($)','Latest Share of Income','Annual Trend ($/yr)','Volatility','Peak Year','Peak Spend ($)'],rows,'FinancialDNAProfilesTable')
        for rr in range(start+1,start+1+len(rows)):
            for c in (8,9,10,12,15): ws.cell(rr,c).number_format=CURRENCY_FORMAT
            for c in (5,11,13): ws.cell(rr,c).number_format=PCT
            ws.cell(rr,2).fill=PatternFill('solid',fgColor={'Rising':LIGHT_RED,'Volatile':LIGHT_YELLOW,'Persistent':LIGHT_BLUE,'Declining':LIGHT_GREEN,'Intermittent':LIGHT_BLUE}.get(str(ws.cell(rr,2).value),LIGHT_BLUE))
        ws.freeze_panes=f'A{start+1}'; self._widths(ws,[24,15,16,16,15,12,12,18,18,18,22,19,14,12,18])
    def _history(self,ws,r):
        rows=[[h.year,h.coverage,h.comparison_eligible,h.level,h.category_id,h.category,float(h.amount),float(h.share_of_income) if h.share_of_income is not None else None,float(h.change_vs_prior) if h.change_vs_prior is not None else None,float(h.change_vs_prior_ratio) if h.change_vs_prior_ratio is not None else None] for h in r.category_history]; self._table(ws,1,1,['Year','Coverage','Comparison Eligible','Level','CategoryID','Category','Amount ($)','Share of Income','Change vs Prior ($)','Change vs Prior (%)'],rows,'CategoryHistoryTable')
        for rr in range(2,2+len(rows)):
            for c in (7,9): ws.cell(rr,c).number_format=CURRENCY_FORMAT
            for c in (8,10): ws.cell(rr,c).number_format=PCT
        ws.freeze_panes='A2'; self._widths(ws,[10,12,20,12,16,28,18,18,22,22])
    def _explorer(self,ws,r,x):
        self._title(ws,'Spending Explorer','Use selectors for a quick answer or filter the table by year, pay period, category and purpose.'); years=sorted({s.year for s in x.sheets}); cats=sorted({str(self.entries[rec.transaction.category_id]['display_name']) for s in x.sheets for rec in s.result.records if self.entries[rec.transaction.category_id]['category_type']!='Income'}); ws['A5']='Selected year'; ws['B5']=r.latest_year; ws['D5']='Selected category'; ws['E5']='All'; ws['G5']='Selected spending'; ws['H5']='=IF($E$5="All",SUMIFS(SpendingExplorerTable[Amount ($)],SpendingExplorerTable[Year],$B$5),SUMIFS(SpendingExplorerTable[Amount ($)],SpendingExplorerTable[Year],$B$5,SpendingExplorerTable[Category],$E$5))'; ws['J5']='Record count'; ws['K5']='=IF($E$5="All",COUNTIFS(SpendingExplorerTable[Year],$B$5),COUNTIFS(SpendingExplorerTable[Year],$B$5,SpendingExplorerTable[Category],$E$5))'
        for c in ('A5','D5','G5','J5'): ws[c].font=Font(bold=True,color=DARK_BLUE); ws[c].fill=PatternFill('solid',fgColor=LIGHT_BLUE)
        for c in ('B5','E5'): ws[c].font=Font(color='0000FF'); ws[c].fill=PatternFill('solid',fgColor='FFF2CC')
        ws['H5'].number_format=CURRENCY_FORMAT; ws['K5'].number_format=INTEGER_FORMAT
        for i,y in enumerate(years,2): ws.cell(i,14,y)
        ws.cell(1,15,'All')
        for i,c in enumerate(cats,2): ws.cell(i,15,c)
        dv1=DataValidation(type='list',formula1=f'=$N$2:$N${1+len(years)}',allow_blank=False); dv2=DataValidation(type='list',formula1=f'=$O$1:$O${1+len(cats)}',allow_blank=False); ws.add_data_validation(dv1); ws.add_data_validation(dv2); dv1.add(ws['B5']); dv2.add(ws['E5']); ws.column_dimensions['N'].hidden=True; ws.column_dimensions['O'].hidden=True
        rows=[]
        for s in x.sheets:
            for rec in s.result.records:
                tx=rec.transaction; e=self.entries[tx.category_id]
                if e['category_type']=='Income':continue
                rows.append([tx.year,tx.period,tx.category_id,e['display_name'],e['master_category'],e['financial_purpose'],e['controllable'],float(tx.amount),tx.source_sheet,tx.source_cell])
        self._table(ws,8,1,['Year','Period','CategoryID','Category','Master Category','Financial Purpose','Controllable','Amount ($)','Source Sheet','Source Cell'],rows,'SpendingExplorerTable')
        for rr in range(9,9+len(rows)): ws.cell(rr,8).number_format=CURRENCY_FORMAT
        ws.freeze_panes='A9'; self._widths(ws,[10,22,14,30,22,22,14,18,14,14,16])
    def _merchants(self,ws,r):
        rows=[[m.original_label,m.mapped_category_id,m.mapped_category,m.suggested_category_id,m.suggested_category,m.rule_id,m.confidence,m.opportunity_group,m.transaction_count,float(m.total_amount),m.latest_year,m.review_status] for m in r.merchants]; self._table(ws,1,1,['Original Label / Merchant','Mapped CategoryID','Mapped Category','Suggested CategoryID','Suggested Category','RuleID','Confidence','Opportunity Group','Occurrences','Historical Amount ($)','Latest Year','Review Status'],rows,'MerchantIntelligenceTable')
        for rr in range(2,2+len(rows)):
            ws.cell(rr,10).number_format=CURRENCY_FORMAT; status=str(ws.cell(rr,12).value); ws.cell(rr,12).fill=PatternFill('solid',fgColor={'Mapped':LIGHT_GREEN,'Suggested — review before mapping':LIGHT_YELLOW,'Unresolved':LIGHT_RED}.get(status,LIGHT_BLUE))
        ws.freeze_panes='A2'; self._widths(ws,[32,18,30,20,30,24,14,28,13,22,12,32])
    def _dashboard(self,ws,r):
        self._section(ws,70,'Decision Intelligence — direct choices and next-step tools',15); top=r.opportunities[0] if r.opportunities else None; combined=sum((o.save_50 for o in r.opportunities),Decimal('0')); base=r.debt_scenarios[0] if r.debt_scenarios else None; focused=next((d for d in r.debt_scenarios if '50%' in d.scenario),None); cards=[('A71:C74','Largest Opportunity',top.opportunity if top else 'N/A',None,LIGHT_YELLOW),('D71:F74','50% Combined Savings',combined,CURRENCY_FORMAT,LIGHT_GREEN),('G71:I74','Current LOC Payoff',base.payoff_months if base else None,'0.0 "months"',LIGHT_BLUE),('J71:L74','Focused LOC Payoff',focused.payoff_months if focused else None,'0.0 "months"',LIGHT_GREEN),('M71:O74','Forecast Scenarios',len(r.forecast_summaries),INTEGER_FORMAT,LIGHT_BLUE)]
        for ref,title,val,fmt,fill in cards:
            a,b=ref.split(':'); minc=ws[a].column; minr=ws[a].row; maxc=ws[b].column; maxr=ws[b].row; ws.merge_cells(start_row=minr,start_column=minc,end_row=minr,end_column=maxc); ws.merge_cells(start_row=minr+1,start_column=minc,end_row=maxr,end_column=maxc); tc=ws.cell(minr,minc,title); tc.font=Font(bold=True,color=DARK_BLUE); tc.fill=PatternFill('solid',fgColor=fill); tc.alignment=Alignment(horizontal='center'); vc=ws.cell(minr+1,minc,val); vc.font=Font(size=14,bold=True,color=BLACK); vc.fill=PatternFill('solid',fgColor=fill); vc.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); vc.number_format=fmt or 'General'
        ws.merge_cells('A76:O76'); ws['A76']=f"Direct example: {top.direct_insight if top else 'No opportunity data available.'} Open Decision_Centre for all scenarios."; ws['A76'].alignment=Alignment(wrap_text=True,vertical='center'); ws['A76'].fill=PatternFill('solid',fgColor=LIGHT_YELLOW); ws['A76'].hyperlink="#'Decision_Centre'!A1"; ws['A76'].style='Hyperlink'; ws.merge_cells('A78:O78'); ws['A78']='Decision Centre  |  Opportunities  |  Debt Planner  |  Forecast  |  Financial DNA  |  Spending Explorer'; ws['A78'].hyperlink="#'Decision_Centre'!A1"; ws['A78'].style='Hyperlink'; ws['A78'].alignment=Alignment(horizontal='center')
