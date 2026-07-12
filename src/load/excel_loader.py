"""Excel loader for validated current-layout imports."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from src.extract import CurrentExtractionResult
from src.transform import CategoryRegistry
from src.validate import ValidationReport

CURRENCY_FORMAT = '$#,##0.00;[Red]($#,##0.00);-'
INTEGER_FORMAT = '#,##0;[Red](#,##0);-'
DATE_TIME_FORMAT = 'yyyy-mm-dd hh:mm:ss'

DARK_BLUE = '17365D'
MID_BLUE = '5B9BD5'
LIGHT_BLUE = 'D9EAF7'
LIGHT_GREEN = 'E2F0D9'
LIGHT_YELLOW = 'FFF2CC'
LIGHT_RED = 'FCE4D6'
WHITE = 'FFFFFF'
BLACK = '000000'
INPUT_BLUE = '0000FF'
SOURCE_GREEN = '008000'


@dataclass(frozen=True, slots=True)
class LoadResult:
    """Summary of a completed FOS workbook load."""

    output_path: Path
    category_rows: int
    transaction_rows: int
    income_rows: int
    exception_rows: int


class ExcelFOSLoader:
    """Write validated normalized records into an FOS Excel workbook."""

    REQUIRED_SHEETS = (
        'Dashboard',
        'Import_Log',
        'DimCategory',
        'FactTransactions',
        'FactIncome',
        'Exceptions',
        'Validation',
    )

    def __init__(self, category_registry: CategoryRegistry) -> None:
        self.category_registry = category_registry

    def load_current(
        self,
        extraction: CurrentExtractionResult,
        validation: ValidationReport,
        output_path: str | Path,
        *,
        source_workbook: str | Path,
        source_sheet: str,
        fos_version: str,
        imported_at: datetime | None = None,
    ) -> LoadResult:
        """Create an FOS workbook from one validated current-layout import."""

        if not validation.is_valid:
            messages = '; '.join(issue.message for issue in validation.errors)
            raise ValueError(f'Cannot load invalid import: {messages}')

        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        imported_at = imported_at or datetime.now(timezone.utc).replace(tzinfo=None)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for sheet_name in self.REQUIRED_SHEETS:
            workbook.create_sheet(sheet_name)

        self._write_categories(workbook['DimCategory'])
        self._write_transactions(workbook['FactTransactions'], extraction)
        self._write_income(workbook['FactIncome'], extraction)
        self._write_exceptions(workbook['Exceptions'], validation)
        self._write_validation(workbook['Validation'], validation)
        self._write_import_log(
            workbook['Import_Log'],
            extraction=extraction,
            validation=validation,
            source_workbook=Path(source_workbook),
            source_sheet=source_sheet,
            fos_version=fos_version,
            imported_at=imported_at,
            output_path=destination,
        )
        self._write_dashboard(
            workbook['Dashboard'],
            source_workbook=Path(source_workbook),
            source_sheet=source_sheet,
            fos_version=fos_version,
            imported_at=imported_at,
        )

        workbook.save(destination)
        workbook.close()

        return LoadResult(
            output_path=destination,
            category_rows=self.category_registry.category_count(),
            transaction_rows=len(extraction.transfers)
            + len(extraction.variable_expenses)
            + len(extraction.fixed_expenses),
            income_rows=len(extraction.income),
            exception_rows=len(validation.exceptions),
        )

    def _write_categories(self, worksheet: Any) -> None:
        headers = [
            'CategoryID',
            'DisplayName',
            'CategoryType',
            'MasterCategory',
            'SubCategory',
            'FinancialPurpose',
            'WeeklyBudget',
            'Controllable',
            'Active',
            'Aliases',
        ]
        rows = []
        for entry in self.category_registry.entries():
            rows.append(
                [
                    entry['category_id'],
                    entry['display_name'],
                    entry['category_type'],
                    entry['master_category'],
                    entry.get('sub_category', ''),
                    entry['financial_purpose'],
                    bool(entry['weekly_budget']),
                    entry['controllable'],
                    bool(entry.get('active', True)),
                    '; '.join(str(alias) for alias in entry.get('aliases', [])),
                ]
            )
        self._write_table(worksheet, headers, rows, 'DimCategoryTable')
        worksheet.freeze_panes = 'A2'
        self._set_widths(worksheet, [14, 30, 18, 20, 22, 22, 14, 14, 10, 48])

    def _write_transactions(
        self, worksheet: Any, extraction: CurrentExtractionResult
    ) -> None:
        headers = [
            'RecordID',
            'Year',
            'TransactionDate',
            'PayPeriod',
            'Section',
            'CategoryID',
            'Description',
            'Amount ($)',
            'SourceSheet',
            'SourceCell',
        ]
        rows = []
        number = 1
        for record in extraction.records:
            if record.section == 'income':
                continue
            transaction = record.transaction
            rows.append(
                [
                    f'TXN-{transaction.year}-{number:04d}',
                    transaction.year,
                    transaction.transaction_date,
                    transaction.period,
                    record.section,
                    transaction.category_id,
                    transaction.description,
                    float(transaction.amount),
                    transaction.source_sheet,
                    transaction.source_cell,
                ]
            )
            number += 1
        self._write_table(worksheet, headers, rows, 'FactTransactionsTable')
        worksheet.freeze_panes = 'A2'
        # The Excel table already owns its AutoFilter. Adding a worksheet-level
        # AutoFilter over the same range creates invalid overlapping filter
        # definitions that Excel repairs by removing the table.
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, 3).number_format = 'yyyy-mm-dd'
            worksheet.cell(row, 8).number_format = CURRENCY_FORMAT
            worksheet.cell(row, 8).font = Font(color=INPUT_BLUE)
            worksheet.cell(row, 9).font = Font(color=SOURCE_GREEN)
            worksheet.cell(row, 10).font = Font(color=SOURCE_GREEN)
        self._set_widths(worksheet, [20, 10, 15, 25, 22, 14, 34, 15, 24, 14])

    def _write_income(self, worksheet: Any, extraction: CurrentExtractionResult) -> None:
        headers = [
            'RecordID',
            'Year',
            'PayPeriod',
            'CategoryID',
            'IncomeSource',
            'Amount ($)',
            'SourceSheet',
            'SourceCell',
        ]
        rows = []
        for number, transaction in enumerate(extraction.income, start=1):
            rows.append(
                [
                    f'INC-{transaction.year}-{number:04d}',
                    transaction.year,
                    transaction.period,
                    transaction.category_id,
                    transaction.description,
                    float(transaction.amount),
                    transaction.source_sheet,
                    transaction.source_cell,
                ]
            )
        self._write_table(worksheet, headers, rows, 'FactIncomeTable')
        worksheet.freeze_panes = 'A2'
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, 6).number_format = CURRENCY_FORMAT
            worksheet.cell(row, 6).font = Font(color=INPUT_BLUE)
            worksheet.cell(row, 7).font = Font(color=SOURCE_GREEN)
            worksheet.cell(row, 8).font = Font(color=SOURCE_GREEN)
        self._set_widths(worksheet, [20, 10, 25, 14, 32, 15, 14, 12])

    def _write_exceptions(self, worksheet: Any, validation: ValidationReport) -> None:
        headers = [
            'NormalizedLabel',
            'SampleLabel',
            'Occurrences',
            'TotalAmount ($)',
            'SourceRefs',
            'ReviewStatus',
            'MappedCategoryID',
            'Notes',
        ]
        rows = [
            [
                exception.normalized_label,
                exception.sample_label,
                exception.occurrences,
                float(exception.total_amount),
                '; '.join(exception.source_refs),
                'Pending',
                '',
                '',
            ]
            for exception in validation.exceptions
        ]
        self._write_table(worksheet, headers, rows, 'ExceptionsTable')
        worksheet.freeze_panes = 'A2'
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, 3).number_format = INTEGER_FORMAT
            worksheet.cell(row, 4).number_format = CURRENCY_FORMAT
            worksheet.cell(row, 6).fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
            worksheet.cell(row, 7).fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
            worksheet.cell(row, 8).fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
        self._set_widths(worksheet, [28, 32, 12, 18, 52, 16, 18, 36])

    def _write_validation(self, worksheet: Any, validation: ValidationReport) -> None:
        worksheet.sheet_view.showGridLines = False
        worksheet['A1'] = 'Validation Summary'
        worksheet['A1'].font = Font(bold=True, color=WHITE, size=14)
        worksheet['A1'].fill = PatternFill('solid', fgColor=DARK_BLUE)
        worksheet.merge_cells('A1:D1')
        worksheet.append(['Metric', 'Value', '', ''])
        for key, value in validation.metrics.items():
            worksheet.append([key, value, '', ''])
        issue_start = worksheet.max_row + 2
        worksheet.cell(issue_start, 1, 'Validation Issues')
        worksheet.cell(issue_start, 1).font = Font(bold=True, color=WHITE)
        worksheet.cell(issue_start, 1).fill = PatternFill('solid', fgColor=DARK_BLUE)
        worksheet.merge_cells(start_row=issue_start, start_column=1, end_row=issue_start, end_column=4)
        worksheet.append(['Code', 'Severity', 'Message', 'SourceRefs'])
        for issue in validation.issues:
            worksheet.append(
                [issue.code, issue.severity, issue.message, '; '.join(issue.source_refs)]
            )
        self._style_header(worksheet, 2, 4)
        self._style_header(worksheet, issue_start + 1, 4)
        self._set_widths(worksheet, [28, 16, 72, 52])
        worksheet.freeze_panes = 'A3'
        for row in range(issue_start + 2, worksheet.max_row + 1):
            severity = str(worksheet.cell(row, 2).value).lower()
            fill = LIGHT_RED if severity == 'error' else LIGHT_YELLOW
            worksheet.cell(row, 2).fill = PatternFill('solid', fgColor=fill)
        worksheet.column_dimensions['C'].width = 72
        worksheet.column_dimensions['D'].width = 52
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    def _write_import_log(
        self,
        worksheet: Any,
        *,
        extraction: CurrentExtractionResult,
        validation: ValidationReport,
        source_workbook: Path,
        source_sheet: str,
        fos_version: str,
        imported_at: datetime,
        output_path: Path,
    ) -> None:
        headers = [
            'ImportDate',
            'SourceWorkbook',
            'SourceSheet',
            'FOSVersion',
            'Status',
            'Periods',
            'SourceRows',
            'NormalizedRecords',
            'UnknownRows',
            'Warnings',
            'Errors',
            'ReconciliationDifference',
            'OutputWorkbook',
        ]
        rows = [
            [
                imported_at,
                source_workbook.name,
                source_sheet,
                fos_version,
                'PASS' if validation.is_valid else 'FAIL',
                len(extraction.periods),
                len(extraction.source_rows),
                len(extraction.records),
                len(extraction.unknown_categories),
                len(validation.warnings),
                len(validation.errors),
                float(Decimal(str(validation.metrics['reconciliation_difference']))),
                output_path.name,
            ]
        ]
        self._write_table(worksheet, headers, rows, 'ImportLogTable')
        worksheet.freeze_panes = 'A2'
        worksheet.cell(2, 1).number_format = DATE_TIME_FORMAT
        worksheet.cell(2, 12).number_format = CURRENCY_FORMAT
        worksheet.cell(2, 5).fill = PatternFill('solid', fgColor=LIGHT_GREEN)
        self._set_widths(
            worksheet,
            [20, 34, 14, 18, 12, 10, 12, 18, 12, 10, 10, 24, 34],
        )

    def _write_dashboard(
        self,
        worksheet: Any,
        *,
        source_workbook: Path,
        source_sheet: str,
        fos_version: str,
        imported_at: datetime,
    ) -> None:
        worksheet.sheet_view.showGridLines = False
        worksheet.merge_cells('A1:H2')
        worksheet['A1'] = 'Family Financial Operating System — Data Load'
        worksheet['A1'].font = Font(size=20, bold=True, color=WHITE)
        worksheet['A1'].fill = PatternFill('solid', fgColor=DARK_BLUE)
        worksheet['A1'].alignment = Alignment(horizontal='left', vertical='center')

        worksheet['A4'] = 'Version'
        worksheet['B4'] = fos_version
        worksheet['D4'] = 'Source'
        worksheet['E4'] = source_workbook.name
        worksheet['G4'] = 'Sheet'
        worksheet['H4'] = source_sheet
        worksheet['A5'] = 'Imported'
        worksheet['B5'] = imported_at
        worksheet['B5'].number_format = DATE_TIME_FORMAT

        for cell in ('A4', 'D4', 'G4', 'A5'):
            worksheet[cell].font = Font(bold=True, color=DARK_BLUE)

        worksheet['A8'] = 'Metric'
        worksheet['B8'] = 'Value'
        metrics = [
            ('Income', '=SUM(FactIncome!F:F)', CURRENCY_FORMAT),
            (
                'Transfers',
                '=SUMIF(FactTransactions!D:D,"transfers",FactTransactions!G:G)',
                CURRENCY_FORMAT,
            ),
            (
                'Variable / irregular expenses',
                '=SUMIF(FactTransactions!D:D,"variable_expenses",FactTransactions!G:G)',
                CURRENCY_FORMAT,
            ),
            (
                'Fixed expenses',
                '=SUMIF(FactTransactions!D:D,"fixed_expenses",FactTransactions!G:G)',
                CURRENCY_FORMAT,
            ),
            ('Unmapped amount', '=SUM(Exceptions!D:D)', CURRENCY_FORMAT),
            ('Normalized records', '=COUNTA(FactTransactions!A:A)-1+COUNTA(FactIncome!A:A)-1', INTEGER_FORMAT),
            ('Exceptions requiring review', '=COUNTA(Exceptions!A:A)-1', INTEGER_FORMAT),
        ]
        for row_number, (label, formula, number_format) in enumerate(metrics, start=9):
            worksheet.cell(row_number, 1, label)
            worksheet.cell(row_number, 2, formula)
            worksheet.cell(row_number, 2).number_format = number_format
            worksheet.cell(row_number, 2).font = Font(color=BLACK)

        self._style_header(worksheet, 8, 2)
        worksheet['D8'] = 'Import Status'
        worksheet['E8'] = 'PASS'
        worksheet['D8'].font = Font(bold=True, color=WHITE)
        worksheet['D8'].fill = PatternFill('solid', fgColor=DARK_BLUE)
        worksheet['E8'].font = Font(bold=True, color=DARK_BLUE)
        worksheet['E8'].fill = PatternFill('solid', fgColor=LIGHT_GREEN)
        worksheet['D10'] = 'Review Exceptions before using this load for analysis.'
        worksheet.merge_cells('D10:H11')
        worksheet['D10'].alignment = Alignment(wrap_text=True, vertical='top')
        worksheet['D10'].fill = PatternFill('solid', fgColor=LIGHT_YELLOW)

        chart = BarChart()
        chart.type = 'col'
        chart.style = 10
        chart.legend = None
        chart.title = '2025 Normalized Cash-flow Sections'
        chart.y_axis.title = 'Amount ($)'
        chart.x_axis.title = 'Section'
        data = Reference(worksheet, min_col=2, min_row=8, max_row=12)
        categories = Reference(worksheet, min_col=1, min_row=9, max_row=12)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 14
        worksheet.add_chart(chart, 'D14')

        self._set_widths(worksheet, [34, 20, 4, 22, 28, 4, 14, 22])
        for row in range(9, 16):
            worksheet.cell(row, 1).border = Border(
                bottom=Side(style='thin', color='D9E1F2')
            )
            worksheet.cell(row, 2).border = Border(
                bottom=Side(style='thin', color='D9E1F2')
            )

    @staticmethod
    def _write_table(
        worksheet: Any,
        headers: list[str],
        rows: list[list[Any]],
        table_name: str,
    ) -> None:
        worksheet.sheet_view.showGridLines = False
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        end_row = max(2, worksheet.max_row)
        if worksheet.max_row == 1:
            worksheet.append([''] * len(headers))
            end_row = 2
        end_column = worksheet.cell(1, len(headers)).column_letter
        table = Table(displayName=table_name, ref=f'A1:{end_column}{end_row}')
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        ExcelFOSLoader._style_header(worksheet, 1, len(headers))

    @staticmethod
    def _style_header(worksheet: Any, row: int, column_count: int) -> None:
        for column in range(1, column_count + 1):
            cell = worksheet.cell(row, column)
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill('solid', fgColor=DARK_BLUE)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    @staticmethod
    def _set_widths(worksheet: Any, widths: list[float]) -> None:
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width
