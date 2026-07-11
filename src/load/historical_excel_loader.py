"""Excel loader for validated workbook-wide historical imports."""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.extract import HistoricalExtractionResult
from src.insights import InsightReport
from src.kpi import AnnualKPI, CurrentSnapshot
from src.load.excel_loader import (
    BLACK,
    CURRENCY_FORMAT,
    DARK_BLUE,
    DATE_TIME_FORMAT,
    INTEGER_FORMAT,
    LIGHT_BLUE,
    LIGHT_GREEN,
    LIGHT_RED,
    LIGHT_YELLOW,
    WHITE,
    ExcelFOSLoader,
    LoadResult,
)
from src.transform import CategoryRegistry
from src.validate import HistoricalValidationReport


class HistoricalExcelFOSLoader(ExcelFOSLoader):
    """Write all official annual worksheets into one FOS workbook."""

    REQUIRED_SHEETS = (
        "Dashboard",
        "Insights",
        "Action_Plan",
        "Spending_Evolution",
        "Import_Log",
        "Annual_KPIs",
        "Current_Snapshot",
        "KPI_Definitions",
        "Insight_Definitions",
        "DimYear",
        "DimCategory",
        "FactTransactions",
        "FactIncome",
        "Exceptions",
        "Validation",
    )

    def __init__(self, category_registry: CategoryRegistry) -> None:
        super().__init__(category_registry)

    def load_historical(
        self,
        extraction: HistoricalExtractionResult,
        validation: HistoricalValidationReport,
        output_path: str | Path,
        *,
        source_workbook: str | Path,
        fos_version: str,
        imported_at: datetime | None = None,
        annual_kpis: tuple[AnnualKPI, ...] = (),
        current_snapshot: CurrentSnapshot | None = None,
        insight_report: InsightReport | None = None,
    ) -> LoadResult:
        if not validation.is_valid:
            messages = "; ".join(issue.message for issue in validation.errors)
            raise ValueError(f"Cannot load invalid historical import: {messages}")

        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        imported_at = imported_at or datetime.now(timezone.utc).replace(tzinfo=None)

        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name in self.REQUIRED_SHEETS:
            workbook.create_sheet(sheet_name)

        self._write_categories(workbook["DimCategory"])
        self._write_transactions(workbook["FactTransactions"], extraction)
        self._write_income(workbook["FactIncome"], extraction)
        self._write_exceptions(workbook["Exceptions"], validation.aggregate)
        self._write_validation(workbook["Validation"], validation.aggregate)
        self._write_dim_year(workbook["DimYear"], extraction, validation)
        self._write_annual_kpis(workbook["Annual_KPIs"], annual_kpis)
        self._write_current_snapshot(workbook["Current_Snapshot"], current_snapshot)
        self._write_kpi_definitions(workbook["KPI_Definitions"])
        self._write_insights(workbook["Insights"], insight_report)
        self._write_action_plan(workbook["Action_Plan"], insight_report)
        self._write_spending_evolution(workbook["Spending_Evolution"], insight_report)
        self._write_insight_definitions(workbook["Insight_Definitions"])
        self._write_import_log_historical(
            workbook["Import_Log"],
            extraction=extraction,
            validation=validation,
            source_workbook=Path(source_workbook),
            fos_version=fos_version,
            imported_at=imported_at,
            output_path=destination,
        )
        self._write_dashboard_historical(
            workbook["Dashboard"],
            extraction=extraction,
            validation=validation,
            source_workbook=Path(source_workbook),
            fos_version=fos_version,
            imported_at=imported_at,
            annual_kpis=annual_kpis,
            current_snapshot=current_snapshot,
            insight_report=insight_report,
        )

        workbook.save(destination)
        workbook.close()

        return LoadResult(
            output_path=destination,
            category_rows=self.category_registry.category_count(),
            transaction_rows=(
                len(extraction.transfers)
                + len(extraction.variable_expenses)
                + len(extraction.fixed_expenses)
            ),
            income_rows=len(extraction.income),
            exception_rows=len(validation.exceptions),
        )

    def _write_dim_year(
        self,
        worksheet: Any,
        extraction: HistoricalExtractionResult,
        validation: HistoricalValidationReport,
    ) -> None:
        reports = {item.sheet_name: item for item in validation.sheets}
        headers = [
            "Year",
            "SourceSheet",
            "Layout",
            "PayPeriods",
            "SourceRows",
            "NormalizedRecords",
            "UnknownRows",
            "Income ($)",
            "Transfers ($)",
            "VariableExpenses ($)",
            "FixedExpenses ($)",
            "UnknownAmount ($)",
            "Status",
        ]
        rows: list[list[Any]] = []
        for sheet in extraction.sheets:
            result = sheet.result
            report = reports[sheet.sheet_name].report
            rows.append(
                [
                    sheet.year,
                    sheet.sheet_name,
                    sheet.layout,
                    len(result.periods),
                    len(result.source_rows),
                    len(result.records),
                    len(result.unknown_categories),
                    float(sum((item.amount for item in result.income), Decimal("0"))),
                    float(sum((item.amount for item in result.transfers), Decimal("0"))),
                    float(
                        sum(
                            (item.amount for item in result.variable_expenses),
                            Decimal("0"),
                        )
                    ),
                    float(
                        sum(
                            (item.amount for item in result.fixed_expenses),
                            Decimal("0"),
                        )
                    ),
                    float(
                        sum(
                            (item.amount for item in result.unknown_categories),
                            Decimal("0"),
                        )
                    ),
                    "PASS" if report.is_valid else "FAIL",
                ]
            )
        self._write_table(worksheet, headers, rows, "DimYearTable")
        worksheet.freeze_panes = "A2"
        for row in range(2, worksheet.max_row + 1):
            for column in range(8, 13):
                worksheet.cell(row, column).number_format = CURRENCY_FORMAT
            worksheet.cell(row, 13).fill = PatternFill(
                "solid", fgColor=LIGHT_GREEN if worksheet.cell(row, 13).value == "PASS" else LIGHT_YELLOW
            )
        self._set_widths(
            worksheet,
            [10, 16, 24, 12, 12, 19, 13, 16, 16, 22, 20, 18, 12],
        )

    def _write_import_log_historical(
        self,
        worksheet: Any,
        *,
        extraction: HistoricalExtractionResult,
        validation: HistoricalValidationReport,
        source_workbook: Path,
        fos_version: str,
        imported_at: datetime,
        output_path: Path,
    ) -> None:
        report_by_sheet = {item.sheet_name: item for item in validation.sheets}
        headers = [
            "ImportDate",
            "SourceWorkbook",
            "SourceSheet",
            "Layout",
            "FOSVersion",
            "Status",
            "Periods",
            "SourceRows",
            "NormalizedRecords",
            "UnknownRows",
            "Warnings",
            "Errors",
            "ReconciliationDifference",
            "OutputWorkbook",
        ]
        rows: list[list[Any]] = []
        for sheet in extraction.sheets:
            report = report_by_sheet[sheet.sheet_name].report
            rows.append(
                [
                    imported_at,
                    source_workbook.name,
                    sheet.sheet_name,
                    sheet.layout,
                    fos_version,
                    "PASS" if report.is_valid else "FAIL",
                    len(sheet.result.periods),
                    len(sheet.result.source_rows),
                    len(sheet.result.records),
                    len(sheet.result.unknown_categories),
                    len(report.warnings),
                    len(report.errors),
                    float(Decimal(str(report.metrics["reconciliation_difference"]))),
                    output_path.name,
                ]
            )
        self._write_table(worksheet, headers, rows, "ImportLogTable")
        worksheet.freeze_panes = "A2"
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, 1).number_format = DATE_TIME_FORMAT
            worksheet.cell(row, 13).number_format = CURRENCY_FORMAT
            worksheet.cell(row, 6).fill = PatternFill(
                "solid", fgColor=LIGHT_GREEN if worksheet.cell(row, 6).value == "PASS" else LIGHT_YELLOW
            )
        self._set_widths(
            worksheet,
            [20, 34, 16, 24, 14, 12, 10, 12, 19, 13, 10, 10, 24, 34],
        )

    def _write_annual_kpis(
        self,
        worksheet: Any,
        annual_kpis: tuple[AnnualKPI, ...],
    ) -> None:
        headers = [
            "Year",
            "Coverage",
            "PayPeriods",
            "TrueIncome ($)",
            "CashFlowAdjustments ($)",
            "FixedExpenses ($)",
            "VariableIrregularExpenses ($)",
            "KnownOperatingExpenses ($)",
            "CoreSpending ($)",
            "LifestyleSpending ($)",
            "WealthBuilding ($)",
            "TargetedDebtReduction ($)",
            "ExcludedTransfers ($)",
            "UnknownAmount ($)",
            "FixedCostRatio",
            "KnownExpenseRatio",
            "WealthBuildingRate",
            "SavingsVelocity",
            "FinancialFlexibility",
            "HousingRatio",
            "TransportationRatio",
            "FoodRatio",
            "LifestyleRatio",
            "DataCoverageRatio",
            "ComparisonEligible",
        ]
        rows: list[list[Any]] = []
        for item in annual_kpis:
            rows.append(
                [
                    item.year,
                    item.coverage_status,
                    item.pay_periods,
                    float(item.true_income),
                    float(item.cash_flow_adjustments),
                    float(item.fixed_expenses),
                    float(item.variable_irregular_expenses),
                    float(item.known_operating_expenses),
                    float(item.core_spending),
                    float(item.lifestyle_spending),
                    float(item.wealth_building),
                    float(item.targeted_debt_reduction),
                    float(item.excluded_transfers),
                    float(item.unknown_amount),
                    float(item.fixed_cost_ratio) if item.fixed_cost_ratio is not None else None,
                    float(item.known_expense_ratio) if item.known_expense_ratio is not None else None,
                    float(item.wealth_building_rate) if item.wealth_building_rate is not None else None,
                    float(item.savings_velocity) if item.savings_velocity is not None else None,
                    float(item.financial_flexibility) if item.financial_flexibility is not None else None,
                    float(item.housing_ratio) if item.housing_ratio is not None else None,
                    float(item.transportation_ratio) if item.transportation_ratio is not None else None,
                    float(item.food_ratio) if item.food_ratio is not None else None,
                    float(item.lifestyle_ratio) if item.lifestyle_ratio is not None else None,
                    float(item.data_coverage_ratio) if item.data_coverage_ratio is not None else None,
                    item.comparison_eligible,
                ]
            )
        self._write_table(worksheet, headers, rows, "AnnualKPIsTable")
        worksheet.freeze_panes = "A2"
        for row in range(2, worksheet.max_row + 1):
            for column in range(4, 15):
                worksheet.cell(row, column).number_format = CURRENCY_FORMAT
            for column in range(15, 25):
                worksheet.cell(row, column).number_format = '0.0%;[Red](0.0%);-'
            if worksheet.cell(row, 2).value == "Partial":
                worksheet.cell(row, 2).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
            if not worksheet.cell(row, 25).value:
                worksheet.cell(row, 25).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        self._set_widths(
            worksheet,
            [10, 12, 12, 16, 22, 18, 28, 26, 18, 20, 20, 25, 20, 18,
             17, 19, 20, 18, 20, 15, 21, 14, 17, 18, 20],
        )
        if len(rows) >= 2:
            chart = LineChart()
            chart.title = "Annual Income, Known Expenses and Wealth Building"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Year"
            chart.style = 10
            data = Reference(worksheet, min_col=4, max_col=11, min_row=1, max_row=worksheet.max_row)
            categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
            # Keep only TrueIncome, KnownOperatingExpenses and WealthBuilding series.
            chart.add_data(Reference(worksheet, min_col=4, min_row=1, max_row=worksheet.max_row), titles_from_data=True)
            chart.add_data(Reference(worksheet, min_col=8, min_row=1, max_row=worksheet.max_row), titles_from_data=True)
            chart.add_data(Reference(worksheet, min_col=11, min_row=1, max_row=worksheet.max_row), titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 15
            worksheet.add_chart(chart, "AA2")

    def _write_current_snapshot(
        self,
        worksheet: Any,
        snapshot: CurrentSnapshot | None,
    ) -> None:
        worksheet.sheet_view.showGridLines = False
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Current Financial Snapshot"
        worksheet["A1"].font = Font(bold=True, color=WHITE, size=14)
        worksheet["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        if snapshot is None:
            worksheet["A3"] = "Current snapshot unavailable."
            return
        rows = [
            ("Source sheet", snapshot.source_sheet, "Source workbook"),
            ("Latest complete cash-flow year", snapshot.latest_complete_year, "Annual_KPIs"),
            ("Total assets", float(snapshot.total_assets), "A & L"),
            ("Total liabilities", float(snapshot.total_liabilities), "A & L"),
            ("Net worth", float(snapshot.net_worth), "A & L"),
            ("Financial assets", float(snapshot.financial_assets), "A & L"),
            ("Savings cash", float(snapshot.savings_cash), "A & L"),
            ("Home equity", float(snapshot.home_equity), "A & L"),
            ("Mortgage balance", float(snapshot.mortgage_balance), "A & L"),
            ("Line of credit balance", float(snapshot.line_of_credit_balance), "A & L"),
            ("Vehicle loan balance", float(snapshot.vehicle_loan_balance), "A & L"),
            (
                "Emergency fund months",
                float(snapshot.emergency_fund_months) if snapshot.emergency_fund_months is not None else None,
                "Savings ÷ monthly core spending",
            ),
            (
                "Unsecured debt ratio",
                float(snapshot.unsecured_debt_ratio) if snapshot.unsecured_debt_ratio is not None else None,
                "LOC ÷ latest complete-year income",
            ),
            (
                "Financial Pressure Index",
                float(snapshot.fpi_score) if snapshot.fpi_score is not None else None,
                "Provisional FPI v1",
            ),
            ("FPI band", snapshot.fpi_band, "Low / Moderate / High / Very High"),
        ]
        worksheet.append(["Metric", "Value", "Source / Definition", "Notes"])
        for label, value, source in rows:
            worksheet.append([label, value, source, ""])
        self._style_header(worksheet, 2, 4)
        for row in range(3, 18):
            label = worksheet.cell(row, 1).value
            if label in {
                "Total assets", "Total liabilities", "Net worth", "Financial assets",
                "Savings cash", "Home equity", "Mortgage balance", "Line of credit balance",
                "Vehicle loan balance",
            }:
                worksheet.cell(row, 2).number_format = CURRENCY_FORMAT
            elif label == "Unsecured debt ratio":
                worksheet.cell(row, 2).number_format = '0.0%;[Red](0.0%);-'
            elif label in {"Emergency fund months", "Financial Pressure Index"}:
                worksheet.cell(row, 2).number_format = '0.0'
        worksheet["D16"] = (
            "FPI is a directional household metric, not an industry-standard score. "
            "It uses the latest complete-year cash flow plus current savings and LOC balances."
        )
        worksheet["D16"].alignment = Alignment(wrap_text=True, vertical="top")
        worksheet["D16"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        self._set_widths(worksheet, [34, 20, 38, 56])
        worksheet.freeze_panes = "A3"

    def _write_kpi_definitions(self, worksheet: Any) -> None:
        headers = ["KPI", "Definition", "Calculation", "Important limitation"]
        rows = [
            ["True Income", "Net household income recorded in the annual sheets.", "INC001 + INC002 + INC003", "Excludes cash-flow adjustment rows."],
            ["Known Operating Expenses", "Mapped fixed, variable, irregular and capital outflows.", "Fixed expenses + variable/irregular expenses", "Unmapped amounts are shown separately."],
            ["Fixed Cost Ratio", "Share of true income committed to mapped fixed expenses.", "Fixed expenses ÷ true income", "Based on recorded workbook classifications."],
            ["Wealth-Building Rate", "Share of income directed to savings, TFSA, RRSP/pension and RESP.", "Wealth-building transfers ÷ true income", "Does not include investment market growth."],
            ["Savings Velocity", "Wealth building plus targeted LOC/student-loan reduction.", "(Wealth building + targeted debt reduction) ÷ true income", "Does not estimate mortgage or vehicle-loan principal."],
            ["Financial Flexibility", "Income remaining after fixed expenses and targeted debt reduction.", "(True income - fixed expenses - targeted debt reduction) ÷ true income", "Variable living costs must still be paid from this amount."],
            ["Data Coverage Ratio", "Share of mapped outflows versus mapped plus unmapped outflows.", "Mapped outflows ÷ (mapped outflows + unknown amount)", "Measures classification completeness, not financial health."],
            ["Comparison Eligible", "Year suitable for direct trend comparison.", "Complete year and data coverage ≥ 85%", "Partial or lower-coverage years remain visible but are flagged."],
            ["FPI v1", "Directional Financial Pressure Index from 0 to 100.", "30% fixed costs + 25% emergency reserve + 30% LOC ratio + 15% known cash margin", "Provisional, household-specific and not an industry-standard score."],
        ]
        self._write_table(worksheet, headers, rows, "KPIDefinitionsTable")
        worksheet.freeze_panes = "A2"
        self._set_widths(worksheet, [26, 55, 58, 62])
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _write_insights(
        self,
        worksheet: Any,
        report: InsightReport | None,
    ) -> None:
        headers = [
            "Rank",
            "Priority",
            "Theme",
            "Headline",
            "Evidence",
            "Implication",
            "Recommended Action",
        ]
        rows: list[list[Any]] = []
        if report is not None:
            rows = [
                [
                    item.rank,
                    item.priority,
                    item.theme,
                    item.headline,
                    item.evidence,
                    item.implication,
                    item.recommended_action,
                ]
                for item in report.insights
            ]
        self._write_table(worksheet, headers, rows, "FinancialInsightsTable")
        worksheet.freeze_panes = "A2"
        self._set_widths(worksheet, [8, 12, 18, 42, 62, 58, 62])
        for row in range(2, worksheet.max_row + 1):
            for column in range(1, 8):
                worksheet.cell(row, column).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
            priority = worksheet.cell(row, 2).value
            worksheet.cell(row, 2).fill = PatternFill(
                "solid", fgColor=self._priority_fill(str(priority))
            )

        worksheet["I1"] = "Insight Summary"
        worksheet["I1"].font = Font(bold=True, color=WHITE, size=12)
        worksheet["I1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        worksheet.merge_cells("I1:K1")
        if report is None:
            worksheet["I3"] = "Insight report unavailable."
            return
        summary = [
            ("Latest complete year", report.latest_year, "year"),
            (
                "Benchmark years",
                ", ".join(str(year) for year in report.benchmark_years) or "N/A",
                "text",
            ),
            ("Monthly core spending", float(report.monthly_core_spending), "currency"),
            (
                "3-month reserve target",
                float(report.emergency_target_amount),
                "currency",
            ),
            ("Reserve funding gap", float(report.emergency_fund_gap), "currency"),
        ]
        for row, (label, value, kind) in enumerate(summary, start=3):
            worksheet.cell(row, 9, label)
            worksheet.cell(row, 10, value)
            worksheet.cell(row, 9).font = Font(bold=True, color=DARK_BLUE)
            worksheet.cell(row, 9).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            worksheet.cell(row, 10).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            if kind == "currency":
                worksheet.cell(row, 10).number_format = CURRENCY_FORMAT
        worksheet.column_dimensions["I"].width = 26
        worksheet.column_dimensions["J"].width = 22
        worksheet.column_dimensions["K"].width = 4

    def _write_action_plan(
        self,
        worksheet: Any,
        report: InsightReport | None,
    ) -> None:
        headers = [
            "Rank",
            "Priority",
            "Area",
            "Action",
            "Why Now",
            "Current",
            "Target",
            "Gap / Headroom",
            "Unit",
            "Measurement",
            "Status",
        ]
        rows: list[list[Any]] = []
        if report is not None:
            rows = [
                [
                    item.rank,
                    item.priority,
                    item.area,
                    item.action,
                    item.rationale,
                    float(item.current_value) if item.current_value is not None else None,
                    float(item.target_value) if item.target_value is not None else None,
                    float(item.gap) if item.gap is not None else None,
                    item.unit,
                    item.measurement,
                    item.status,
                ]
                for item in report.actions
            ]
        self._write_table(worksheet, headers, rows, "ActionPlanTable")
        worksheet.freeze_panes = "A2"
        self._set_widths(
            worksheet,
            [8, 12, 18, 48, 55, 16, 16, 18, 14, 42, 16],
        )
        for row in range(2, worksheet.max_row + 1):
            for column in range(1, 12):
                worksheet.cell(row, column).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
            for column in (6, 7, 8):
                worksheet.cell(row, column).number_format = CURRENCY_FORMAT
            priority = worksheet.cell(row, 2).value
            worksheet.cell(row, 2).fill = PatternFill(
                "solid", fgColor=self._priority_fill(str(priority))
            )
            status = str(worksheet.cell(row, 11).value)
            status_fill = {
                "Complete": LIGHT_GREEN,
                "On track": LIGHT_GREEN,
                "In progress": LIGHT_BLUE,
                "Monitor": LIGHT_YELLOW,
                "Review": LIGHT_YELLOW,
                "Needs action": LIGHT_RED,
            }.get(status, LIGHT_BLUE)
            worksheet.cell(row, 11).fill = PatternFill("solid", fgColor=status_fill)

    def _write_spending_evolution(
        self,
        worksheet: Any,
        report: InsightReport | None,
    ) -> None:
        headers = [
            "MetricID",
            "Group",
            "Metric",
            "LatestYear",
            "LatestValue ($)",
            "PriorYear",
            "PriorValue ($)",
            "ChangeVsPrior ($)",
            "ChangeVsPrior (%)",
            "BenchmarkYears",
            "BenchmarkAverage ($)",
            "ChangeVsBenchmark ($)",
            "ChangeVsBenchmark (%)",
            "Direction",
            "Signal",
        ]
        rows: list[list[Any]] = []
        if report is not None:
            for item in report.spending_evolution:
                rows.append(
                    [
                        item.metric_id,
                        item.metric_group,
                        item.metric,
                        item.latest_year,
                        float(item.latest_value),
                        item.prior_year,
                        float(item.prior_value) if item.prior_value is not None else None,
                        float(item.change_vs_prior_amount)
                        if item.change_vs_prior_amount is not None
                        else None,
                        float(item.change_vs_prior_ratio)
                        if item.change_vs_prior_ratio is not None
                        else None,
                        ", ".join(str(year) for year in item.benchmark_years),
                        float(item.benchmark_average)
                        if item.benchmark_average is not None
                        else None,
                        float(item.change_vs_benchmark_amount)
                        if item.change_vs_benchmark_amount is not None
                        else None,
                        float(item.change_vs_benchmark_ratio)
                        if item.change_vs_benchmark_ratio is not None
                        else None,
                        item.direction,
                        item.signal,
                    ]
                )
        self._write_table(worksheet, headers, rows, "SpendingEvolutionTable")
        worksheet.freeze_panes = "A2"
        for row in range(2, worksheet.max_row + 1):
            for column in (5, 7, 8, 11, 12):
                worksheet.cell(row, column).number_format = CURRENCY_FORMAT
            for column in (9, 13):
                worksheet.cell(row, column).number_format = '0.0%;[Red](0.0%);-'
            signal = str(worksheet.cell(row, 15).value)
            worksheet.cell(row, 15).fill = PatternFill(
                "solid",
                fgColor={
                    "Favourable": LIGHT_GREEN,
                    "Stable": LIGHT_BLUE,
                    "Watch": LIGHT_YELLOW,
                    "Context": LIGHT_BLUE,
                }.get(signal, LIGHT_BLUE),
            )
        self._set_widths(
            worksheet,
            [28, 20, 30, 12, 18, 12, 18, 20, 20, 20, 22, 24, 24, 14, 14],
        )

    def _write_insight_definitions(self, worksheet: Any) -> None:
        headers = ["Rule", "FOS v0.6.0 definition", "Use", "Limitation"]
        rows = [
            [
                "Benchmark window",
                "Up to the three comparison-eligible years immediately before the latest complete year.",
                "Provides recent context without treating partial years as equivalent.",
                "A short benchmark can be affected by unusual one-time years.",
            ],
            [
                "Three-month reserve target",
                "Latest complete-year core spending ÷ 12 × 3.",
                "Creates a measurable liquidity target.",
                "An internal FOS operating rule, not an external financial standard.",
            ],
            [
                "Fixed-cost watch level",
                "30% of latest complete-year true income.",
                "Flags reduced flexibility before the FPI fixed-cost pressure begins.",
                "Category mapping and household circumstances affect interpretation.",
            ],
            [
                "Wealth-building floor",
                "10% of latest complete-year true income.",
                "Protects long-term saving momentum while other actions are addressed.",
                "Does not include investment growth or employer-only pension value.",
            ],
            [
                "Largest spending increases",
                "Mapped non-transfer master categories ranked by dollar increase versus the prior comparison year.",
                "Directs transaction review to the largest changes first.",
                "An increase is not automatically wasteful or avoidable.",
            ],
            [
                "Action-plan gaps",
                "Current value compared with the stated FOS target or watch level.",
                "Turns insights into measurable operating actions.",
                "Gap values are directional and do not replace a detailed financial plan.",
            ],
        ]
        self._write_table(worksheet, headers, rows, "InsightDefinitionsTable")
        worksheet.freeze_panes = "A2"
        self._set_widths(worksheet, [28, 64, 58, 64])
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _priority_fill(priority: str) -> str:
        return {
            "High": LIGHT_RED,
            "Medium": LIGHT_YELLOW,
            "Low": LIGHT_GREEN,
            "Protect": LIGHT_GREEN,
        }.get(priority, LIGHT_BLUE)

    def _write_dashboard_historical(
        self,
        worksheet: Any,
        *,
        extraction: HistoricalExtractionResult,
        validation: HistoricalValidationReport,
        source_workbook: Path,
        fos_version: str,
        imported_at: datetime,
        annual_kpis: tuple[AnnualKPI, ...],
        current_snapshot: CurrentSnapshot | None,
        insight_report: InsightReport | None,
    ) -> None:
        """Create the v0.6.0 executive dashboard.

        The dashboard intentionally displays calculated values rather than adding
        another calculation layer. Hidden helper columns provide compact,
        chart-safe datasets and keep all chart references inside the dashboard.
        """

        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.zoomScale = 85
        worksheet.freeze_panes = "A7"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.sheet_properties.outlinePr.summaryBelow = True

        # Title and metadata.
        worksheet.merge_cells("A1:O2")
        worksheet["A1"] = "Family Financial Operating System — Executive Dashboard"
        worksheet["A1"].font = Font(size=20, bold=True, color=WHITE)
        worksheet["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

        worksheet.merge_cells("A3:O3")
        worksheet["A3"] = (
            "Current household position, latest complete-year performance and "
            "comparison-eligible historical trends"
        )
        worksheet["A3"].font = Font(italic=True, color="44546A")
        worksheet["A3"].alignment = Alignment(vertical="center")

        metadata = [
            ("A4", "Version", "B4", fos_version),
            ("D4", "Source", "E4", source_workbook.name),
            ("J4", "Years imported", "K4", len(extraction.sheets)),
            ("M4", "Validation", "N4", "PASS" if validation.is_valid else "FAIL"),
            ("A5", "Imported", "B5", imported_at),
            (
                "D5",
                "Source range",
                "E5",
                (
                    f"{extraction.sheets[0].year}–{extraction.sheets[-1].year}"
                    if extraction.sheets
                    else ""
                ),
            ),
            ("J5", "Unmapped labels", "K5", len(validation.exceptions)),
            ("M5", "Official sheets", "N5", len(extraction.sheets)),
        ]
        for label_cell, label, value_cell, value in metadata:
            worksheet[label_cell] = label
            worksheet[label_cell].font = Font(bold=True, color=DARK_BLUE)
            worksheet[value_cell] = value
        worksheet["B5"].number_format = DATE_TIME_FORMAT
        worksheet["N4"].font = Font(bold=True, color=DARK_BLUE)
        worksheet["N4"].fill = PatternFill(
            "solid", fgColor=LIGHT_GREEN if validation.is_valid else LIGHT_RED
        )

        complete = [item for item in annual_kpis if item.coverage_status == "Complete"]
        latest = max(complete, key=lambda item: item.year) if complete else None
        trend_kpis = [item for item in annual_kpis if item.comparison_eligible]
        if not trend_kpis:
            trend_kpis = complete or list(annual_kpis)
        trend_kpis = sorted(trend_kpis, key=lambda item: item.year)

        # Current-position cards.
        self._write_dashboard_section(worksheet, 7, "Current Position")
        total_debt = current_snapshot.total_liabilities if current_snapshot else None
        current_cards = [
            ("A8:C11", "Net Worth", current_snapshot.net_worth if current_snapshot else None, CURRENCY_FORMAT, LIGHT_BLUE),
            ("D8:F11", "Liquid Savings", current_snapshot.savings_cash if current_snapshot else None, CURRENCY_FORMAT, LIGHT_GREEN),
            ("G8:I11", "Total Debt", total_debt, CURRENCY_FORMAT, LIGHT_YELLOW),
            ("J8:L11", "Emergency Fund", current_snapshot.emergency_fund_months if current_snapshot else None, '0.0 "months"', LIGHT_YELLOW),
            ("M8:O11", "Financial Pressure Index", current_snapshot.fpi_score if current_snapshot else None, '0.0', self._fpi_fill(current_snapshot.fpi_band if current_snapshot else "Unavailable")),
        ]
        for ref, title, value, number_format, fill in current_cards:
            self._write_dashboard_card(
                worksheet,
                ref,
                title,
                value,
                number_format,
                fill,
                subtitle=(current_snapshot.fpi_band if title == "Financial Pressure Index" and current_snapshot else None),
            )

        # Latest-year performance cards.
        year_label = latest.year if latest else "N/A"
        self._write_dashboard_section(
            worksheet, 13, f"Latest Complete-Year Performance ({year_label})"
        )
        latest_cards = [
            ("A14:C17", "True Income", latest.true_income if latest else None, CURRENCY_FORMAT, LIGHT_BLUE),
            ("D14:F17", "Known Operating Expenses", latest.known_operating_expenses if latest else None, CURRENCY_FORMAT, LIGHT_YELLOW),
            ("G14:I17", "Wealth Building", latest.wealth_building if latest else None, CURRENCY_FORMAT, LIGHT_GREEN),
            ("J14:L17", "Savings Velocity", latest.savings_velocity if latest else None, '0.0%;[Red](0.0%);-', LIGHT_GREEN),
            ("M14:O17", "Financial Flexibility", latest.financial_flexibility if latest else None, '0.0%;[Red](0.0%);-', LIGHT_BLUE),
        ]
        for ref, title, value, number_format, fill in latest_cards:
            self._write_dashboard_card(
                worksheet, ref, title, value, number_format, fill
            )

        # Status strip.
        latest_unknown = latest.unknown_amount if latest else None
        status_items = [
            ("A19:C20", "Validation", "PASS" if validation.is_valid else "FAIL", LIGHT_GREEN if validation.is_valid else LIGHT_RED),
            ("D19:F20", "Latest Complete Year", latest.year if latest else "N/A", LIGHT_BLUE),
            ("G19:I20", "Data Coverage", latest.data_coverage_ratio if latest else None, LIGHT_GREEN),
            ("J19:L20", "Unmapped Amount", latest_unknown, LIGHT_YELLOW),
            ("M19:O20", "Trend Eligibility", "Eligible" if latest and latest.comparison_eligible else "Review", LIGHT_GREEN if latest and latest.comparison_eligible else LIGHT_YELLOW),
        ]
        for ref, title, value, fill in status_items:
            number_format = "General"
            if title == "Data Coverage":
                number_format = '0.0%;[Red](0.0%);-'
            elif title == "Unmapped Amount":
                number_format = CURRENCY_FORMAT
            self._write_dashboard_status(worksheet, ref, title, value, number_format, fill)

        # Hidden chart datasets.
        self._write_dashboard_chart_data(
            worksheet,
            trend_kpis=trend_kpis,
            latest=latest,
            current_snapshot=current_snapshot,
        )

        # Charts.
        cash_chart = LineChart()
        cash_chart.title = "Annual Cash Flow Trend"
        cash_chart.y_axis.title = "Amount ($)"
        cash_chart.x_axis.title = "Year"
        cash_chart.style = 10
        cash_chart.height = 8.5
        cash_chart.width = 15.5
        if trend_kpis:
            end = 1 + len(trend_kpis)
            cash_chart.add_data(
                Reference(worksheet, min_col=17, max_col=19, min_row=1, max_row=end),
                titles_from_data=True,
            )
            cash_chart.set_categories(
                Reference(worksheet, min_col=16, min_row=2, max_row=end)
            )
            for series, title in zip(
                cash_chart.series,
                ("True Income", "Known Expenses", "Wealth Building"),
            ):
                series.tx = SeriesLabel(v=title)
        cash_chart.legend.position = "b"
        worksheet.add_chart(cash_chart, "A23")

        ratio_chart = LineChart()
        ratio_chart.title = "Key Ratio Trend"
        ratio_chart.y_axis.title = "Percent of Income"
        ratio_chart.x_axis.title = "Year"
        ratio_chart.y_axis.number_format = "0%"
        ratio_chart.style = 13
        ratio_chart.height = 8.5
        ratio_chart.width = 13.5
        if trend_kpis:
            end = 1 + len(trend_kpis)
            ratio_chart.add_data(
                Reference(worksheet, min_col=20, max_col=22, min_row=1, max_row=end),
                titles_from_data=True,
            )
            ratio_chart.set_categories(
                Reference(worksheet, min_col=16, min_row=2, max_row=end)
            )
            for series, title in zip(
                ratio_chart.series,
                ("Fixed Cost", "Wealth Building", "Financial Flexibility"),
            ):
                series.tx = SeriesLabel(v=title)
        ratio_chart.legend.position = "b"
        worksheet.add_chart(ratio_chart, "I23")

        balance_chart = BarChart()
        balance_chart.type = "col"
        balance_chart.title = "Current Balance Sheet"
        balance_chart.y_axis.title = "Amount ($)"
        balance_chart.style = 10
        balance_chart.height = 8.5
        balance_chart.width = 15.5
        balance_chart.add_data(
            Reference(worksheet, min_col=25, min_row=1, max_row=4),
            titles_from_data=True,
        )
        balance_chart.set_categories(
            Reference(worksheet, min_col=24, min_row=2, max_row=4)
        )
        balance_chart.legend = None
        worksheet.add_chart(balance_chart, "A40")

        mix_chart = DoughnutChart()
        mix_chart.title = f"{year_label} Known Spending Mix"
        mix_chart.style = 10
        mix_chart.height = 8.5
        mix_chart.width = 13.5
        mix_chart.add_data(
            Reference(worksheet, min_col=28, min_row=1, max_row=6),
            titles_from_data=True,
        )
        mix_chart.set_categories(
            Reference(worksheet, min_col=27, min_row=2, max_row=6)
        )
        mix_chart.holeSize = 55
        mix_chart.dataLabels = DataLabelList()
        mix_chart.dataLabels.showPercent = True
        mix_chart.legend.position = "r"
        worksheet.add_chart(mix_chart, "I40")

        worksheet.merge_cells("A57:O59")
        worksheet["A57"] = (
            "Dashboard notes: trend charts use comparison-eligible years where available. "
            "Partial years remain in Annual_KPIs but are excluded from direct trend interpretation. "
            "FPI is a directional household metric, not an industry-standard score. Review "
            "Exceptions before treating unmapped amounts as fully classified."
        )
        worksheet["A57"].alignment = Alignment(wrap_text=True, vertical="top")
        worksheet["A57"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        worksheet["A57"].font = Font(color="7F6000")

        self._write_dashboard_section(worksheet, 61, "Executive Takeaways")
        if insight_report is not None:
            for index, insight in enumerate(insight_report.insights[:3]):
                start_row = 62 + index * 2
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=start_row,
                    end_column=3,
                )
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=4,
                    end_row=start_row,
                    end_column=15,
                )
                worksheet.merge_cells(
                    start_row=start_row + 1,
                    start_column=1,
                    end_row=start_row + 1,
                    end_column=15,
                )
                label = worksheet.cell(start_row, 1)
                label.value = f"#{insight.rank} {insight.priority} — {insight.theme}"
                label.font = Font(bold=True, color=DARK_BLUE)
                label.fill = PatternFill(
                    "solid", fgColor=self._priority_fill(insight.priority)
                )
                headline = worksheet.cell(start_row, 4)
                headline.value = insight.headline
                headline.font = Font(bold=True, color=BLACK)
                headline.fill = PatternFill(
                    "solid", fgColor=self._priority_fill(insight.priority)
                )
                detail = worksheet.cell(start_row + 1, 1)
                detail.value = insight.evidence + " Action: " + insight.recommended_action
                detail.alignment = Alignment(wrap_text=True, vertical="top")
                detail.fill = PatternFill("solid", fgColor="F2F2F2")
            worksheet.merge_cells("A68:O68")
            worksheet["A68"] = "Open Insights and Action_Plan for the complete evidence and measurable targets."
            worksheet["A68"].hyperlink = "#'Insights'!A1"
            worksheet["A68"].style = "Hyperlink"
            worksheet["A68"].alignment = Alignment(horizontal="center")
        else:
            worksheet.merge_cells("A62:O63")
            worksheet["A62"] = "Insight report unavailable."

        # Balanced dashboard widths. Helper columns P:AB remain hidden.
        dashboard_widths = [12, 12, 12, 12, 16, 12, 12, 12, 12, 15, 12, 12, 14, 12, 12]
        for column, width in enumerate(dashboard_widths, start=1):
            worksheet.column_dimensions[get_column_letter(column)].width = width
        for column in range(16, 29):
            worksheet.column_dimensions[get_column_letter(column)].hidden = True
        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[2].height = 28
        worksheet.row_dimensions[3].height = 22
        for row in (7, 13):
            worksheet.row_dimensions[row].height = 22
        for row in range(8, 18):
            worksheet.row_dimensions[row].height = 22
        for row in (57, 58, 59):
            worksheet.row_dimensions[row].height = 22
        worksheet.row_dimensions[61].height = 22
        for row in range(62, 68):
            worksheet.row_dimensions[row].height = 30

    @staticmethod
    def _write_dashboard_section(worksheet: Any, row: int, title: str) -> None:
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        cell = worksheet.cell(row, 1)
        cell.value = title
        cell.font = Font(bold=True, color=WHITE, size=12)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def _write_dashboard_card(
        worksheet: Any,
        ref: str,
        title: str,
        value: Any,
        number_format: str,
        fill: str,
        *,
        subtitle: str | None = None,
    ) -> None:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(ref)
        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=min_row,
            end_column=max_col,
        )
        worksheet.merge_cells(
            start_row=min_row + 1,
            start_column=min_col,
            end_row=max_row - (1 if subtitle else 0),
            end_column=max_col,
        )
        label = worksheet.cell(min_row, min_col)
        label.value = title
        label.font = Font(bold=True, color=DARK_BLUE, size=10)
        label.fill = PatternFill("solid", fgColor=fill)
        label.alignment = Alignment(horizontal="center", vertical="center")

        value_cell = worksheet.cell(min_row + 1, min_col)
        value_cell.value = float(value) if isinstance(value, Decimal) else value
        value_cell.number_format = number_format
        value_cell.font = Font(bold=True, color=BLACK, size=17)
        value_cell.fill = PatternFill("solid", fgColor=fill)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

        if subtitle:
            worksheet.merge_cells(
                start_row=max_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )
            sub = worksheet.cell(max_row, min_col)
            sub.value = subtitle
            sub.font = Font(italic=True, color="44546A", size=9)
            sub.fill = PatternFill("solid", fgColor=fill)
            sub.alignment = Alignment(horizontal="center", vertical="center")

        thin = Side(style="thin", color="B4C6E7")
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row, col)
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.border = Border(
                    left=thin if col == min_col else Side(style=None),
                    right=thin if col == max_col else Side(style=None),
                    top=thin if row == min_row else Side(style=None),
                    bottom=thin if row == max_row else Side(style=None),
                )

    @staticmethod
    def _write_dashboard_status(
        worksheet: Any,
        ref: str,
        title: str,
        value: Any,
        number_format: str,
        fill: str,
    ) -> None:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(ref)
        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=min_row,
            end_column=max_col,
        )
        worksheet.merge_cells(
            start_row=max_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )
        label = worksheet.cell(min_row, min_col)
        label.value = title
        label.font = Font(bold=True, color=DARK_BLUE, size=9)
        label.fill = PatternFill("solid", fgColor=fill)
        label.alignment = Alignment(horizontal="center", vertical="center")

        metric = worksheet.cell(max_row, min_col)
        metric.value = float(value) if isinstance(value, Decimal) else value
        metric.number_format = number_format
        metric.font = Font(bold=True, color=BLACK, size=11)
        metric.fill = PatternFill("solid", fgColor=fill)
        metric.alignment = Alignment(horizontal="center", vertical="center")

        thin = Side(style="thin", color="B4C6E7")
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row, col)
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.border = Border(
                    left=thin if col == min_col else Side(style=None),
                    right=thin if col == max_col else Side(style=None),
                    top=thin if row == min_row else Side(style=None),
                    bottom=thin if row == max_row else Side(style=None),
                )

    @staticmethod
    def _fpi_fill(band: str) -> str:
        return {
            "Low": LIGHT_GREEN,
            "Moderate": LIGHT_YELLOW,
            "High": "F8CBAD",
            "Very High": LIGHT_RED,
        }.get(band, LIGHT_BLUE)

    @staticmethod
    def _write_dashboard_chart_data(
        worksheet: Any,
        *,
        trend_kpis: list[AnnualKPI],
        latest: AnnualKPI | None,
        current_snapshot: CurrentSnapshot | None,
    ) -> None:
        trend_headers = [
            "Year",
            "True Income",
            "Known Expenses",
            "Wealth Building",
            "Fixed Cost Ratio",
            "Wealth-Building Rate",
            "Financial Flexibility",
        ]
        for col, header in enumerate(trend_headers, start=16):
            worksheet.cell(1, col, header)
        for row, item in enumerate(trend_kpis, start=2):
            values = [
                item.year,
                float(item.true_income),
                float(item.known_operating_expenses),
                float(item.wealth_building),
                float(item.fixed_cost_ratio) if item.fixed_cost_ratio is not None else None,
                float(item.wealth_building_rate) if item.wealth_building_rate is not None else None,
                float(item.financial_flexibility) if item.financial_flexibility is not None else None,
            ]
            for col, value in enumerate(values, start=16):
                worksheet.cell(row, col, value)

        worksheet["X1"] = "Balance Sheet"
        worksheet["Y1"] = "Amount"
        balance_values = [
            ("Assets", current_snapshot.total_assets if current_snapshot else Decimal("0")),
            ("Liabilities", current_snapshot.total_liabilities if current_snapshot else Decimal("0")),
            ("Net Worth", current_snapshot.net_worth if current_snapshot else Decimal("0")),
        ]
        for row, (label, value) in enumerate(balance_values, start=2):
            worksheet.cell(row, 24, label)
            worksheet.cell(row, 25, float(value))

        worksheet["AA1"] = "Spending Category"
        worksheet["AB1"] = "Amount"
        if latest:
            housing = latest.true_income * (latest.housing_ratio or Decimal("0"))
            transportation = latest.true_income * (latest.transportation_ratio or Decimal("0"))
            food = latest.true_income * (latest.food_ratio or Decimal("0"))
            lifestyle = latest.lifestyle_spending
            other = max(
                Decimal("0"),
                latest.known_operating_expenses
                - housing
                - transportation
                - food
                - lifestyle,
            )
        else:
            housing = transportation = food = lifestyle = other = Decimal("0")
        mix_values = [
            ("Housing", housing),
            ("Transportation", transportation),
            ("Food", food),
            ("Lifestyle", lifestyle),
            ("Other Known", other),
        ]
        for row, (label, value) in enumerate(mix_values, start=2):
            worksheet.cell(row, 27, label)
            worksheet.cell(row, 28, float(value))

