"""Verify the complete FOS v0.6.0 insight engine."""

from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.extract import CurrentLayoutExtractor, LayoutDetector  # noqa: E402
from src.models import Category, ImportResult, ImportSession, Transaction  # noqa: E402,F401
from src.load import ExcelFOSLoader  # noqa: E402
from src.pipeline import CurrentYearPipeline  # noqa: E402
from src.historical_pipeline import HistoricalPipeline  # noqa: E402
from src.extract import HistoricalWorkbookExtractor  # noqa: E402
from src.validate import HistoricalImportValidator  # noqa: E402
from src.transform import CategoryRegistry  # noqa: E402
from src.kpi import KPIEngine  # noqa: E402
from src.insights import InsightsEngine  # noqa: E402
from src.validate import ImportValidator, write_validation_report  # noqa: E402


def workbook_sheet_names(workbook_path: Path) -> list[str]:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    try:
        with ZipFile(workbook_path) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
    except (FileNotFoundError, BadZipFile, KeyError) as exc:
        raise ValueError(f"Unable to read workbook structure: {workbook_path}") from exc

    root = ElementTree.fromstring(workbook_xml)
    sheets = root.find("main:sheets", namespace)
    if sheets is None:
        return []
    return [sheet.attrib["name"] for sheet in sheets]


def verify_current_snapshot_consistency(snapshot) -> None:
    """Validate snapshot relationships without hardcoding mutable balances."""
    expected_net_worth = snapshot.total_assets - snapshot.total_liabilities
    if snapshot.net_worth != expected_net_worth:
        raise ValueError(
            "Current net worth calculation mismatch: "
            f"expected {expected_net_worth}, found {snapshot.net_worth}."
        )
    if snapshot.fpi_score is not None:
        if snapshot.fpi_score < Decimal("0") or snapshot.fpi_score > Decimal("100"):
            raise ValueError("Current FPI score is outside the valid 0–100 range.")
        expected_band = KPIEngine._fpi_band(snapshot.fpi_score)
        if snapshot.fpi_band != expected_band:
            raise ValueError(
                "Current FPI band mismatch: "
                f"expected {expected_band}, found {snapshot.fpi_band}."
            )


def verify_current_snapshot_sheet(worksheet, snapshot) -> None:
    """Confirm the generated Current_Snapshot sheet matches the calculated snapshot."""
    cell_net_worth = Decimal(str(worksheet["B7"].value))
    if cell_net_worth != snapshot.net_worth:
        raise ValueError(
            "Current net worth KPI output mismatch: "
            f"expected {snapshot.net_worth}, found {cell_net_worth}."
        )

    cell_fpi = worksheet["B16"].value
    if snapshot.fpi_score is None:
        if cell_fpi is not None:
            raise ValueError("Current FPI KPI output should be blank.")
    elif Decimal(str(cell_fpi)) != snapshot.fpi_score:
        raise ValueError(
            "Current FPI KPI output mismatch: "
            f"expected {snapshot.fpi_score}, found {cell_fpi}."
        )

    if worksheet["B17"].value != snapshot.fpi_band:
        raise ValueError(
            "Current FPI band output mismatch: "
            f"expected {snapshot.fpi_band}, found {worksheet['B17'].value}."
        )


def verify_required_workbook_aliases(workbook_path: Path, registry: CategoryRegistry) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        annual_pattern = re.compile(r"20\d{2}( \(old\))?")
        labels: set[str] = set()
        for sheet_name in workbook.sheetnames:
            if not annual_pattern.fullmatch(sheet_name):
                continue
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        labels.add(" ".join(cell.value.split()))
    finally:
        workbook.close()

    required = {
        "Costco": "FOD001",
        "Mortgage (21st)": "HOU001",
        "Fuel": "TRA002",
        "Visa payment": "TRF005",
        "Questrade TFSA": "TRF002",
        "Canadian Tire": "HOU012",
    }
    for label, expected_id in required.items():
        if label not in labels:
            raise ValueError(f"Required workbook label not found: {label}")
        actual_id = registry.lookup(label).category_id
        if actual_id != expected_id:
            raise ValueError(
                f"Workbook label '{label}' mapped to '{actual_id}', expected '{expected_id}'."
            )


def verify_2025_extraction_validation_and_load(
    workbook_path: Path, registry: CategoryRegistry
) -> None:
    extractor = CurrentLayoutExtractor(registry)
    result = extractor.extract(workbook_path, "2025")
    report = ImportValidator(registry).validate_current(result)

    expected = {
        "periods": 26,
        "source_rows": 497,
        "records": 463,
        "unknowns": 34,
        "unknown_labels": 28,
        "income_count": 85,
        "transfer_count": 86,
        "variable_count": 157,
        "fixed_count": 135,
        "income_total": Decimal("215878.10"),
        "transfer_total": Decimal("119717.02"),
        "variable_total": Decimal("28492.48"),
        "fixed_total": Decimal("59879.19"),
        "unknown_total": Decimal("8112.69"),
        "source_total": Decimal("432079.48"),
    }
    actual = {
        "periods": len(result.periods),
        "source_rows": len(result.source_rows),
        "records": len(result.records),
        "unknowns": len(result.unknown_categories),
        "unknown_labels": len(report.exceptions),
        "income_count": len(result.income),
        "transfer_count": len(result.transfers),
        "variable_count": len(result.variable_expenses),
        "fixed_count": len(result.fixed_expenses),
        "income_total": sum((item.amount for item in result.income), Decimal("0")),
        "transfer_total": sum((item.amount for item in result.transfers), Decimal("0")),
        "variable_total": sum(
            (item.amount for item in result.variable_expenses), Decimal("0")
        ),
        "fixed_total": sum((item.amount for item in result.fixed_expenses), Decimal("0")),
        "unknown_total": sum(
            (item.amount for item in result.unknown_categories), Decimal("0")
        ),
        "source_total": sum((item.amount for item in result.source_rows), Decimal("0")),
    }
    if actual != expected:
        differences = [
            f"{key}: expected {expected[key]}, found {actual[key]}"
            for key in expected
            if actual[key] != expected[key]
        ]
        raise ValueError("2025 extraction mismatch: " + "; ".join(differences))

    if not report.is_valid:
        raise ValueError(
            "2025 validation failed: "
            + "; ".join(issue.message for issue in report.errors)
        )
    if Decimal(str(report.metrics["reconciliation_difference"])) != 0:
        raise ValueError("2025 source reconciliation did not equal zero.")

    with TemporaryDirectory() as temporary_dir:
        summary_path, exceptions_path = write_validation_report(report, temporary_dir)
        if not summary_path.is_file() or not exceptions_path.is_file():
            raise ValueError("Validation report files were not created.")

    print(f"- 2025 pay periods extracted: {actual['periods']}")
    print(f"- 2025 source financial rows: {actual['source_rows']}")
    print(f"- 2025 normalized records: {actual['records']}")
    print(f"- 2025 unmapped records: {actual['unknowns']}")
    print(f"- 2025 unmapped labels: {actual['unknown_labels']}")
    print(f"- 2025 source total: ${actual['source_total']:,.2f}")
    print(f"- 2025 reconciliation difference: {report.metrics['reconciliation_difference']}")
    with TemporaryDirectory() as temporary_dir:
        output_path = Path(temporary_dir) / "Financial_Operating_System.xlsx"
        load_result = ExcelFOSLoader(registry).load_current(
            result,
            report,
            output_path,
            source_workbook=workbook_path,
            source_sheet="2025",
            fos_version="0.6.0",
        )
        from openpyxl import load_workbook

        output_workbook = load_workbook(output_path, data_only=False, read_only=False)
        try:
            expected_sheets = list(ExcelFOSLoader.REQUIRED_SHEETS)
            if output_workbook.sheetnames != expected_sheets:
                raise ValueError(
                    f"FOS workbook sheets mismatch: {output_workbook.sheetnames}"
                )
            if output_workbook["FactTransactions"].max_row != 379:
                raise ValueError("Expected 378 transaction rows plus header.")
            if output_workbook["FactIncome"].max_row != 86:
                raise ValueError("Expected 85 income rows plus header.")
            if output_workbook["Exceptions"].max_row != 29:
                raise ValueError("Expected 28 exception rows plus header.")
            if output_workbook["Import_Log"]["E2"].value != "PASS":
                raise ValueError("Import log did not record PASS status.")
            if output_workbook["Dashboard"]["B9"].value != "=SUM(FactIncome!F:F)":
                raise ValueError("Dashboard income formula is missing.")
        finally:
            output_workbook.close()

        if load_result.transaction_rows != 378 or load_result.income_rows != 85:
            raise ValueError("FOS loader row counts did not match the verified import.")

    with TemporaryDirectory() as temporary_dir:
        integrated_output = Path(temporary_dir) / "Integrated_FOS.xlsx"
        pipeline_result = CurrentYearPipeline(PROJECT_ROOT).run(
            workbook_path,
            sheet_name="2025",
            output_path=integrated_output,
            fos_version="0.6.0",
        )
        if not integrated_output.is_file():
            raise ValueError("Integrated pipeline did not create the FOS workbook.")
        if pipeline_result.load_result.transaction_rows != 378:
            raise ValueError("Integrated pipeline transaction count mismatch.")
        if not pipeline_result.validation_summary_path.is_file():
            raise ValueError("Integrated pipeline validation summary is missing.")
        if not pipeline_result.exceptions_path.is_file():
            raise ValueError("Integrated pipeline exceptions report is missing.")

    print("- 2025 validation and exceptions reports: OK")
    print("- 2025 FOS Excel workbook load: OK")
    print("- End-to-end update pipeline: OK")



def verify_historical_import(workbook_path: Path, registry: CategoryRegistry) -> None:
    detector = LayoutDetector(PROJECT_ROOT / "config" / "layouts.yaml")
    extraction = HistoricalWorkbookExtractor(detector, registry).extract(workbook_path)
    validation = HistoricalImportValidator(ImportValidator(registry)).validate(extraction)

    expected = {
        "sheets": 18,
        "excluded": 1,
        "periods": 412,
        "source_rows": 6430,
        "records": 5655,
        "unknowns": 775,
        "unknown_labels": 603,
        "income_count": 911,
        "transaction_count": 4744,
        "income_total": Decimal("2247082.59260000000275"),
        "transfer_total": Decimal("1604663.809"),
        "variable_total": Decimal("213060.070000000000014"),
        "fixed_total": Decimal("463401.31"),
        "unknown_total": Decimal("223723.82"),
        "source_total": Decimal("4751931.601600000002764"),
    }
    actual = {
        "sheets": len(extraction.sheets),
        "excluded": len(extraction.excluded_sheets),
        "periods": sum(len(sheet.result.periods) for sheet in extraction.sheets),
        "source_rows": len(extraction.source_rows),
        "records": len(extraction.records),
        "unknowns": len(extraction.unknown_categories),
        "unknown_labels": len(validation.exceptions),
        "income_count": len(extraction.income),
        "transaction_count": (
            len(extraction.transfers)
            + len(extraction.variable_expenses)
            + len(extraction.fixed_expenses)
        ),
        "income_total": sum((item.amount for item in extraction.income), Decimal("0")),
        "transfer_total": sum((item.amount for item in extraction.transfers), Decimal("0")),
        "variable_total": sum(
            (item.amount for item in extraction.variable_expenses), Decimal("0")
        ),
        "fixed_total": sum(
            (item.amount for item in extraction.fixed_expenses), Decimal("0")
        ),
        "unknown_total": sum(
            (item.amount for item in extraction.unknown_categories), Decimal("0")
        ),
        "source_total": sum(
            (item.amount for item in extraction.source_rows), Decimal("0")
        ),
    }
    if actual != expected:
        differences = [
            f"{key}: expected {expected[key]}, found {actual[key]}"
            for key in expected
            if actual[key] != expected[key]
        ]
        raise ValueError("Historical extraction mismatch: " + "; ".join(differences))
    if not validation.is_valid:
        raise ValueError(
            "Historical validation failed: "
            + "; ".join(issue.message for issue in validation.errors)
        )
    if Decimal(str(validation.metrics["reconciliation_difference"])) != 0:
        raise ValueError("Historical source reconciliation did not equal zero.")
    if "2017 (old)" not in extraction.excluded_sheets:
        raise ValueError("Archived 2017 (old) sheet was not excluded.")

    kpi_engine = KPIEngine(registry)
    annual_kpis = kpi_engine.calculate_annual(extraction)
    snapshot = kpi_engine.calculate_current_snapshot(workbook_path, annual_kpis)
    latest = next(item for item in annual_kpis if item.year == 2025)
    if latest.wealth_building_rate.quantize(Decimal("0.0001")) != Decimal("0.1089"):
        raise ValueError("2025 wealth-building rate mismatch.")
    if latest.financial_flexibility.quantize(Decimal("0.0001")) != Decimal("0.7224"):
        raise ValueError("2025 financial-flexibility KPI mismatch.")
    verify_current_snapshot_consistency(snapshot)
    insight_report = InsightsEngine(registry).analyze(
        extraction, annual_kpis, snapshot
    )
    if insight_report.latest_year != 2025:
        raise ValueError("Insight report did not select 2025 as the latest complete year.")
    if insight_report.benchmark_years != (2022, 2023, 2024):
        raise ValueError("Insight benchmark years mismatch.")
    expected_reserve_target = latest.core_spending / Decimal("12") * Decimal("3")
    expected_reserve_gap = max(Decimal("0"), expected_reserve_target - snapshot.savings_cash)
    if insight_report.emergency_target_amount != expected_reserve_target:
        raise ValueError("Three-month emergency target mismatch.")
    if insight_report.emergency_fund_gap != expected_reserve_gap:
        raise ValueError("Emergency-fund gap mismatch.")
    if len(insight_report.insights) != 6 or len(insight_report.actions) != 6:
        raise ValueError("Expected six insights and six action-plan items.")

    with TemporaryDirectory() as temporary_dir:
        output = Path(temporary_dir) / "Historical_FOS.xlsx"
        pipeline = HistoricalPipeline(PROJECT_ROOT).run(
            workbook_path,
            output_path=output,
            fos_version="0.6.0",
        )
        if not output.is_file():
            raise ValueError("Historical pipeline did not create the FOS workbook.")
        from openpyxl import load_workbook

        workbook = load_workbook(output, data_only=False, read_only=False)
        try:
            if workbook["DimYear"].max_row != 19:
                raise ValueError("Expected 18 historical year rows plus header.")
            if workbook["FactTransactions"].max_row != 4745:
                raise ValueError("Historical transaction row count mismatch.")
            if workbook["FactIncome"].max_row != 912:
                raise ValueError("Historical income row count mismatch.")
            if workbook["Exceptions"].max_row != 604:
                raise ValueError("Historical exception row count mismatch.")
            if workbook["FactTransactions"].auto_filter.ref is not None:
                raise ValueError("FactTransactions contains an overlapping AutoFilter.")
            if "FactTransactionsTable" not in workbook["FactTransactions"].tables:
                raise ValueError("FactTransactionsTable is missing.")
            if workbook["Annual_KPIs"].max_row != 19:
                raise ValueError("Expected 18 annual KPI rows plus header.")
            if workbook["KPI_Definitions"].max_row != 10:
                raise ValueError("KPI definitions are missing.")
            if workbook["Insights"].max_row != 7:
                raise ValueError("Expected six financial insights plus header.")
            if workbook["Action_Plan"].max_row != 7:
                raise ValueError("Expected six action-plan items plus header.")
            if workbook["Spending_Evolution"].max_row < 10:
                raise ValueError("Spending evolution output is incomplete.")
            if workbook["Insight_Definitions"].max_row != 7:
                raise ValueError("Insight definitions are missing.")
            verify_current_snapshot_sheet(workbook["Current_Snapshot"], snapshot)
            dashboard = workbook["Dashboard"]
            if dashboard["A1"].value != "Family Financial Operating System — Executive Dashboard":
                raise ValueError("Executive dashboard title is missing.")
            if dashboard["A8"].value != "Net Worth":
                raise ValueError("Executive dashboard current-position cards are missing.")
            if dashboard["A14"].value != "True Income":
                raise ValueError("Executive dashboard latest-year cards are missing.")
            if len(dashboard._charts) != 4:
                raise ValueError("Executive dashboard must contain four charts.")
            if not dashboard.column_dimensions["P"].hidden or not dashboard.column_dimensions["AB"].hidden:
                raise ValueError("Executive dashboard helper columns must remain hidden.")
            if dashboard.auto_filter.ref is not None:
                raise ValueError("Executive dashboard contains an unexpected AutoFilter.")
            if dashboard["A61"].value != "Executive Takeaways":
                raise ValueError("Executive dashboard insight summary is missing.")
        finally:
            workbook.close()
        if pipeline.load_result.transaction_rows != 4744:
            raise ValueError("Historical pipeline transaction count mismatch.")
        if not pipeline.validation_summary_path.is_file():
            raise ValueError("Historical validation summary is missing.")
        if not pipeline.exceptions_path.is_file():
            raise ValueError("Historical exceptions report is missing.")
        if pipeline.insight_report is None:
            raise ValueError("Historical pipeline insight report is missing.")

    print(f"- Historical worksheets imported: {actual['sheets']}")
    print(f"- Historical pay periods extracted: {actual['periods']}")
    print(f"- Historical normalized records: {actual['records']}")
    print(f"- Historical unmapped records: {actual['unknowns']}")
    print(f"- Historical source total: ${actual['source_total']:,.2f}")
    print("- Archived 2017 (old) exclusion: OK")
    print("- Historical reconciliation difference: 0")
    print("- Historical FOS workbook load: OK")
    print("- Executive dashboard cards and charts: OK")
    print("- Spending evolution, financial insights and action plan: OK")
    print(f"- Three-month reserve target: ${insight_report.emergency_target_amount:,.2f}")
    print(f"- Reserve funding gap: ${insight_report.emergency_fund_gap:,.2f}")
    print("- 2025 wealth-building rate: 10.9%")
    print("- 2025 financial flexibility: 72.2%")
    print(f"- Current net worth: ${snapshot.net_worth:,.2f}")
    if snapshot.fpi_score is None:
        print("- Provisional FPI: unavailable")
    else:
        print(f"- Provisional FPI: {snapshot.fpi_score} ({snapshot.fpi_band})")

def main() -> int:
    parser = argparse.ArgumentParser(description="Verify FOS v0.6.0")
    parser.add_argument("--workbook", type=Path, help="Optional private Budget workbook path.")
    args = parser.parse_args()

    detector = LayoutDetector(PROJECT_ROOT / "config" / "layouts.yaml")
    registry = CategoryRegistry(PROJECT_ROOT / "config" / "categories.yaml")

    print("FOS v0.6.0 verification")
    print("- Core models: OK")
    print(f"- Configured categories: {registry.category_count()}")
    print(f"- Configured aliases: {registry.alias_count()}")
    print(f"- Costco mapping: {registry.lookup('Costco').display_name}")
    print(f"- Canadian Tire mapping: {registry.lookup('Canadian Tire').display_name}")
    print(f"- Mortgage (21st) mapping: {registry.lookup('Mortgage (21st)').display_name}")
    print(f"- 2010 layout: {detector.detect('2010')}")
    print(f"- 2017 (old) layout: {detector.detect('2017 (old)')}")
    print(f"- 2025 layout: {detector.detect('2025')}")

    if args.workbook:
        source_sheets = workbook_sheet_names(args.workbook)
        configured = set(detector.configured_sheets())
        annual_pattern = re.compile(r"20\d{2}( \(old\))?")
        annual_sheets = {name for name in source_sheets if annual_pattern.fullmatch(name)}
        missing = sorted(annual_sheets - configured)
        extra = sorted(configured - annual_sheets)
        print(f"- Workbook annual sheets found: {len(annual_sheets)}")
        if missing:
            print(f"- Unconfigured annual sheets: {', '.join(missing)}")
            return 1
        if extra:
            print(f"- Configured sheets absent from workbook: {', '.join(extra)}")
            return 1
        print("- Workbook/layout configuration match: OK")
        verify_required_workbook_aliases(args.workbook, registry)
        print("- Workbook/category dictionary checks: OK")
        verify_2025_extraction_validation_and_load(args.workbook, registry)
        verify_historical_import(args.workbook, registry)

    print("Verification PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
