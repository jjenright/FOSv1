from dataclasses import replace
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src.extract import HistoricalWorkbookExtractor, LayoutDetector
from src.kpi import KPIEngine
from src.transform import CategoryRegistry
from tests.test_historical_pipeline import build_mixed_history_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def add_balance_sheet(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.create_sheet("A & L")
    sheet.append(["Item", "Asset or Liability", "Amount", "Comments"])
    sheet.append(["House", "Asset", 500000, "test"])
    sheet.append(["Mortgage", "Liability", 250000, "test"])
    sheet.append(["Savings", "Asset", 12000, "test"])
    sheet.append(["TFSA 1", "Asset", 30000, "test"])
    sheet.append(["Credit Line", "Liability", 5000, "test"])
    sheet.append(["Truck loan", "Liability", 10000, "test"])
    workbook.save(path)
    workbook.close()


def build_engine_and_history(source: Path):
    detector = LayoutDetector(PROJECT_ROOT / "config" / "layouts.yaml")
    registry = CategoryRegistry(PROJECT_ROOT / "config" / "categories.yaml")
    extraction = HistoricalWorkbookExtractor(detector, registry).extract(
        source, sheets=("2010", "2017")
    )
    return registry, KPIEngine(registry), extraction


def test_annual_kpis_classify_complete_and_partial_years(tmp_path) -> None:
    source = tmp_path / "Budget.xlsx"
    build_mixed_history_workbook(source)
    _registry, engine, extraction = build_engine_and_history(source)

    annual = engine.calculate_annual(extraction)

    assert len(annual) == 2
    assert annual[0].year == 2010
    assert annual[0].coverage_status == "Partial"
    assert annual[0].true_income == Decimal("2000")
    assert annual[1].year == 2017
    assert annual[1].known_operating_expenses == Decimal("1500")


def test_current_snapshot_calculates_net_worth_and_fpi(tmp_path) -> None:
    source = tmp_path / "Budget.xlsx"
    build_mixed_history_workbook(source)
    add_balance_sheet(source)
    _registry, engine, extraction = build_engine_and_history(source)
    annual = list(engine.calculate_annual(extraction))
    annual[-1] = replace(annual[-1], coverage_status="Complete")

    snapshot = engine.calculate_current_snapshot(source, tuple(annual))

    assert snapshot.total_assets == Decimal("542000")
    assert snapshot.total_liabilities == Decimal("265000")
    assert snapshot.net_worth == Decimal("277000")
    assert snapshot.home_equity == Decimal("250000")
    assert snapshot.financial_assets == Decimal("42000")
    assert snapshot.fpi_score is not None
    assert snapshot.fpi_band in {"Low", "Moderate", "High", "Very High"}


def test_fpi_bands_are_stable() -> None:
    assert KPIEngine._fpi_band(Decimal("0")) == "Low"
    assert KPIEngine._fpi_band(Decimal("25")) == "Moderate"
    assert KPIEngine._fpi_band(Decimal("50")) == "High"
    assert KPIEngine._fpi_band(Decimal("75")) == "Very High"
