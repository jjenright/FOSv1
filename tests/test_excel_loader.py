from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.extract import CurrentLayoutExtractor
from src.load import ExcelFOSLoader
from src.transform import CategoryRegistry
from src.validate import ImportValidator
from tests.test_current_extractor import build_current_sheet

CONFIG = Path(__file__).resolve().parents[1] / 'config' / 'categories.yaml'


def test_loads_validated_records_into_fos_workbook(tmp_path) -> None:
    source_workbook, worksheet = build_current_sheet()
    registry = CategoryRegistry(CONFIG)
    extraction = CurrentLayoutExtractor(registry).extract_worksheet(worksheet)
    validation = ImportValidator(registry).validate_current(extraction)
    output = tmp_path / 'fos.xlsx'

    result = ExcelFOSLoader(registry).load_current(
        extraction,
        validation,
        output,
        source_workbook='Budget.xlsx',
        source_sheet='2025',
        fos_version='test',
        imported_at=datetime(2026, 7, 11, 12, 0, 0),
    )
    source_workbook.close()

    assert output.is_file()
    assert result.category_rows == registry.category_count()
    assert result.transaction_rows == 3
    assert result.income_rows == 1
    assert result.exception_rows == 1

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == list(ExcelFOSLoader.REQUIRED_SHEETS)
        assert workbook['FactTransactions'].max_row == 4
        assert workbook['FactIncome'].max_row == 2
        assert workbook['Exceptions'].max_row == 2
        assert workbook['Import_Log']['E2'].value == 'PASS'
        assert workbook['Dashboard']['B9'].value == '=SUM(FactIncome!F:F)'
        assert workbook['FactTransactions']['H2'].value == '2025'
        assert workbook['FactTransactions']['I2'].value in {'C2', 'E2'}
    finally:
        workbook.close()


def test_rejects_invalid_validation_report(tmp_path) -> None:
    source_workbook, worksheet = build_current_sheet()
    registry = CategoryRegistry(CONFIG)
    extraction = CurrentLayoutExtractor(registry).extract_worksheet(worksheet)

    from src.extract import CurrentExtractionResult, SourceRow
    from decimal import Decimal

    invalid_extraction = CurrentExtractionResult(
        records=(),
        unknown_categories=(),
        source_rows=(SourceRow('P1', 'Costco', Decimal('10'), '2025', 'C2'),),
        periods=('P1',),
    )
    validation = ImportValidator(registry).validate_current(invalid_extraction)
    source_workbook.close()

    with pytest.raises(ValueError, match='Cannot load invalid import'):
        ExcelFOSLoader(registry).load_current(
            extraction,
            validation,
            tmp_path / 'fos.xlsx',
            source_workbook='Budget.xlsx',
            source_sheet='2025',
            fos_version='test',
        )
