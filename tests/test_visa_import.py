from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from src.extract import HistoricalWorkbookExtractor, LayoutDetector, VisaSheetExtractor
from src.transform import CategoryRegistry


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def add_visa_sheet(workbook: Workbook, title: str = "Visa Transactions 2025"):
    sheet = workbook.create_sheet(title)
    sheet.append(
        [
            "TransactionID",
            "TransactionDate",
            "PostingDate",
            "ActivityDescription",
            "NormalizedMerchant",
            "Amount ($)",
            "TransactionType",
            "CategoryID",
            "SuggestedCategory",
            "MasterCategory",
            "CategoryConfidence",
            "ReviewRequired",
            "SourceScreenshot",
            "SourceRow",
        ]
    )
    sheet.append(
        [
            "VISA-001",
            datetime(2025, 6, 10),
            datetime(2025, 6, 11),
            "LCBO #123",
            "LCBO",
            125.50,
            "Purchase",
            "FOD003",
            "Alcohol",
            "Food",
            "High",
            "No",
            "screen1.png",
            1,
        ]
    )
    sheet.append(
        [
            "VISA-002",
            datetime(2025, 6, 12),
            datetime(2025, 6, 12),
            "PAYMENT - THANK YOU",
            "Credit Card Payment",
            -500,
            "Transfer",
            "TRF005",
            "Credit Card Payment",
            "Transfer",
            "High",
            "No",
            "screen1.png",
            2,
        ]
    )
    sheet.append(
        [
            "VISA-003",
            datetime(2025, 6, 15),
            datetime(2025, 6, 16),
            "VILLAGE PIZZA",
            "Village Pizza",
            80.25,
            "Purchase",
            "FOD002",
            "Dining Out",
            "Food",
            "High",
            "No",
            "screen1.png",
            3,
        ]
    )
    return sheet


def test_visa_sheet_extractor_imports_purchases_and_skips_payments() -> None:
    registry = CategoryRegistry(PROJECT_ROOT / "config" / "categories.yaml")
    workbook = Workbook()
    workbook.active.title = "Notes"
    add_visa_sheet(workbook)

    result = VisaSheetExtractor(registry).extract_workbook_sheets(workbook)

    assert tuple(result.by_year) == (2025,)
    year = result.by_year[2025]
    assert len(year.records) == 2
    assert len(year.source_rows) == 2
    assert len(year.unknown_categories) == 0
    assert year.records[0].transaction.description == "LCBO"
    assert year.records[0].transaction.transaction_date.isoformat() == "2025-06-10"
    assert result.stats.imported_records == 2
    assert result.stats.skipped_transfers == 1
    assert result.stats.imported_total == Decimal("205.75")


def test_historical_extractor_merges_visa_rows_into_matching_year(tmp_path) -> None:
    source = tmp_path / "Budget.xlsx"
    workbook = Workbook()
    annual = workbook.active
    annual.title = "2025"
    annual.append(["Jan 1 - Jan 14", "PayAmount", "Bi-weekly", "BWAmount", "Monthlies", "MonAmount"])
    annual.append(["JE's Pay", 2500, "Costco", 200, "Visa payment", 500])
    add_visa_sheet(workbook)
    workbook.save(source)
    workbook.close()

    detector = LayoutDetector(PROJECT_ROOT / "config" / "layouts.yaml")
    registry = CategoryRegistry(PROJECT_ROOT / "config" / "categories.yaml")
    extraction = HistoricalWorkbookExtractor(detector, registry).extract(
        source, sheets=("2025",)
    )

    assert len(extraction.sheets) == 1
    result = extraction.sheets[0].result
    assert extraction.sheets[0].layout.endswith("+visa")
    assert len(result.periods) == 1  # Visa months do not distort pay-period coverage.
    assert len(result.records) == 5  # pay, Costco, Visa transfer, LCBO, pizza
    assert len(result.source_rows) == 5
    assert extraction.visa_import is not None
    assert extraction.visa_import.imported_records == 2
    assert extraction.visa_import.skipped_transfers == 1
