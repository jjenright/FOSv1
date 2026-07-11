from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.extract import HistoricalWorkbookExtractor, LayoutDetector
from src.historical_pipeline import HistoricalPipeline
from src.transform import CategoryRegistry


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def build_mixed_history_workbook(path: Path) -> None:
    workbook = Workbook()
    legacy = workbook.active
    legacy.title = "2010"
    legacy["A1"] = "Pay Period"
    legacy["B1"] = "Jan 1 - Jan 14"
    legacy["A2"] = "Pay"
    legacy["B2"] = 2000
    legacy["A4"] = "Bi-Weekly Deductions"
    legacy["B5"] = 100
    legacy["C5"] = "Gas"

    current = workbook.create_sheet("2017")
    current["A1"] = "Jul 12 - Jul 25"
    current["B1"] = "PayAmount"
    current["C1"] = "Bi-weekly"
    current["D1"] = "BWAmount"
    current["E1"] = "Monthlies"
    current["F1"] = "MonAmount"
    current["A2"] = "JE's Pay"
    current["B2"] = 2500
    current["C2"] = "Costco"
    current["D2"] = 200
    current["E2"] = "Mortgage (21st)"
    current["F2"] = 1300

    archived = workbook.create_sheet("2017 (old)")
    archived["A1"] = "Pay Period"
    archived["B1"] = "Jan 1 - Jan 14"
    archived["A2"] = "Pay - JE"
    archived["B2"] = 9999
    workbook.save(path)
    workbook.close()


def test_official_history_excludes_archived_2017() -> None:
    detector = LayoutDetector(PROJECT_ROOT / "config" / "layouts.yaml")
    registry = CategoryRegistry(PROJECT_ROOT / "config" / "categories.yaml")
    extractor = HistoricalWorkbookExtractor(detector, registry)

    assert "2017" in extractor.official_sheets()
    assert "2017 (old)" not in extractor.official_sheets()


def test_historical_pipeline_loads_mixed_layouts(tmp_path) -> None:
    source = tmp_path / "Budget.xlsx"
    output = tmp_path / "Financial_Operating_System.xlsx"
    build_mixed_history_workbook(source)

    result = HistoricalPipeline(PROJECT_ROOT).run(
        source,
        output_path=output,
        sheets=("2010", "2017"),
        fos_version="test",
    )

    assert result.validation.is_valid
    assert result.load_result.transaction_rows == 3
    assert result.load_result.income_rows == 2
    assert output.is_file()

    workbook = load_workbook(output, data_only=False, read_only=False)
    try:
        assert workbook.sheetnames[0] == "Dashboard"
        assert workbook["DimYear"].max_row == 3
        assert workbook["Import_Log"].max_row == 3
        assert workbook["FactTransactions"].max_row == 4
        assert workbook["FactIncome"].max_row == 3
        assert workbook["Dashboard"]["B16"].value == "=COUNTA(DimYear!A:A)-1"
    finally:
        workbook.close()
